"""
Conversion API Module

Provides programmatic access to the EPUB conversion pipeline.
This module can be used by the API server or called directly.

Jobs are stored in MongoDB for persistence across container restarts.
ISBN is used as the job_id for meaningful tracking.
Output files are stored in the storage backend (GridFS, S3, or Local).
"""

import json
import logging
import os
import re
import sys
import threading
from dataclasses import asdict, dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

# Ensure parent directory is in path for imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from conversion_tracker import ConversionStatus, ConversionTracker

# Thread-safe context for concurrent conversions
try:
    from thread_safe_context import (
        thread_safe_conversion, ConversionContext,
        get_authority_or_global, get_mapper_or_global
    )
    THREAD_SAFE_CONTEXT_AVAILABLE = True
except ImportError:
    THREAD_SAFE_CONTEXT_AVAILABLE = False

logger = logging.getLogger(__name__)

# Maximum length for job IDs (ISBN-based identifiers)
MAX_JOB_ID_LENGTH = 24

# MongoDB integration
try:
    from src.db.mongodb_client import MONGODB_AVAILABLE, get_mongodb_client
except ImportError:
    MONGODB_AVAILABLE = False
    logger.warning("MongoDB client not available - using in-memory fallback")

# Storage integration
try:
    from src.storage import get_storage, StorageBackend
    STORAGE_AVAILABLE = True
except ImportError:
    STORAGE_AVAILABLE = False
    logger.warning("Storage module not available - files will only be saved locally")


def extract_isbn_from_filename(filename: str, max_length: int = MAX_JOB_ID_LENGTH) -> str:
    """
    Extract ISBN from filename.

    Looks for patterns like:
    - 9781234567890.epub
    - 978-1-234-56789-0.epub
    - book_9781234567890.epub

    Falls back to filename stem if no ISBN found.
    All IDs are limited to max_length characters (default 24).
    """
    # Remove extension and path
    name = Path(filename).stem

    # Try to find ISBN-13 pattern (13 digits, optionally with hyphens)
    isbn_pattern = r'(?:97[89][-\s]?(?:\d[-\s]?){9}\d|\d{13})'
    match = re.search(isbn_pattern, name.replace('-', '').replace(' ', ''))

    if match:
        # Clean up - remove hyphens/spaces
        isbn = re.sub(r'[-\s]', '', match.group(0))
        return isbn[:max_length]

    # Try ISBN-10 pattern
    isbn10_pattern = r'\d{9}[\dXx]'
    match = re.search(isbn10_pattern, name.replace('-', '').replace(' ', ''))
    if match:
        return match.group(0).upper()[:max_length]

    # Fallback to filename stem (cleaned and truncated)
    cleaned = re.sub(r'[^\w\d]', '_', name)
    return cleaned[:max_length]


class JobStatus(str, Enum):
    """Status of a conversion job."""
    PENDING = "pending"
    RUNNING = "running"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"


@dataclass
class ConversionJob:
    """Represents a conversion job."""
    job_id: str
    input_file: str
    output_dir: str
    status: JobStatus = JobStatus.PENDING
    progress: int = 0
    message: str = ""
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    result: Optional[Dict[str, Any]] = None
    error: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return asdict(self)


class ConversionAPI:
    """
    API for managing EPUB conversions.

    This class provides methods for:
    - Starting conversion jobs (using ISBN as job_id)
    - Checking job status (from MongoDB)
    - Getting conversion results
    - Accessing dashboard data

    Jobs are persisted in MongoDB for access across container restarts.
    """

    # MongoDB collection name for jobs
    JOBS_COLLECTION = "conversion_jobs"

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize the Conversion API.

        Args:
            output_dir: Default output directory for conversions
        """
        self.output_dir = output_dir or Path(__file__).parent.parent / "Output"
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # In-memory cache for active jobs (running threads)
        self._active_jobs: Dict[str, ConversionJob] = {}
        self._job_threads: Dict[str, threading.Thread] = {}
        self._lock = threading.Lock()

        # Progress callbacks for real-time updates
        self._progress_callbacks: Dict[str, List[Callable]] = {}

        # MongoDB client
        self._mongodb_client = None
        if MONGODB_AVAILABLE:
            try:
                self._mongodb_client = get_mongodb_client()
                if self._mongodb_client.connect():
                    logger.info("ConversionAPI connected to MongoDB")
                else:
                    logger.warning("ConversionAPI MongoDB connection failed - using fallback")
                    self._mongodb_client = None
            except Exception as e:
                logger.warning(f"MongoDB initialization failed: {e} - using fallback")
                self._mongodb_client = None

        logger.info(f"ConversionAPI initialized with output_dir: {self.output_dir}")

    def _save_job_to_mongodb(self, job: ConversionJob) -> bool:
        """Save or update a job in MongoDB."""
        if not self._mongodb_client:
            return False

        try:
            db = self._mongodb_client._db
            collection = db[self.JOBS_COLLECTION]

            job_data = job.to_dict()
            job_data['_id'] = job.job_id  # Use ISBN as document ID
            job_data['updated_at'] = datetime.now().isoformat()

            # Upsert - update if exists, insert if not
            collection.replace_one(
                {'_id': job.job_id},
                job_data,
                upsert=True
            )
            return True
        except Exception as e:
            logger.warning(f"Failed to save job to MongoDB: {e}")
            return False

    def _get_job_from_mongodb(self, job_id: str) -> Optional[ConversionJob]:
        """Retrieve a job from MongoDB by job_id (ISBN)."""
        if not self._mongodb_client:
            return None

        try:
            db = self._mongodb_client._db
            collection = db[self.JOBS_COLLECTION]

            doc = collection.find_one({'_id': job_id})
            if doc:
                # Convert MongoDB doc to ConversionJob
                doc.pop('_id', None)
                doc.pop('updated_at', None)
                return ConversionJob(**doc)
            return None
        except Exception as e:
            logger.warning(f"Failed to get job from MongoDB: {e}")
            return None

    def _list_jobs_from_mongodb(
        self,
        status: Optional[str] = None,
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """List jobs from MongoDB with optional filtering."""
        if not self._mongodb_client:
            return []

        try:
            db = self._mongodb_client._db
            collection = db[self.JOBS_COLLECTION]

            query = {}
            if status:
                query['status'] = status

            cursor = collection.find(query).sort('created_at', -1).limit(limit)

            jobs = []
            for doc in cursor:
                doc['job_id'] = doc.pop('_id', doc.get('job_id'))
                doc.pop('updated_at', None)
                jobs.append(doc)
            return jobs
        except Exception as e:
            logger.warning(f"Failed to list jobs from MongoDB: {e}")
            return []

    def get_job(self, job_id: str) -> Optional[ConversionJob]:
        """
        Get a job by ID, checking active jobs first, then MongoDB.
        """
        # Check active (in-memory) jobs first
        with self._lock:
            if job_id in self._active_jobs:
                return self._active_jobs[job_id]

        # Check MongoDB
        return self._get_job_from_mongodb(job_id)

    def start_conversion(
        self,
        input_file: str,
        output_dir: Optional[str] = None,
        async_mode: bool = True,
        debug: bool = False,
        job_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Start an EPUB conversion job.

        Args:
            input_file: Path to the input EPUB file
            output_dir: Optional output directory (uses default if not specified)
            async_mode: If True, run conversion in background thread
            debug: Enable debug logging
            job_id: Optional job ID (defaults to ISBN extracted from filename)

        Returns:
            Dict with job_id (ISBN) and status information
        """
        input_path = Path(input_file).resolve()

        # Validate input file
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

        if input_path.suffix.lower() not in ['.epub', '.epub3']:
            raise ValueError(f"Invalid file type. Expected .epub or .epub3, got {input_path.suffix}")

        # Determine output directory
        out_dir = Path(output_dir).resolve() if output_dir else self.output_dir
        out_dir.mkdir(parents=True, exist_ok=True)

        # Use ISBN as job_id (extracted from filename if not provided)
        if not job_id:
            job_id = extract_isbn_from_filename(input_path.name)
        else:
            # Ensure explicitly provided job_id is within length limit
            job_id = job_id[:MAX_JOB_ID_LENGTH]

        # Check if job already exists and is running
        existing_job = self.get_job(job_id)
        if existing_job and existing_job.status == JobStatus.RUNNING:
            return {
                "job_id": job_id,
                "status": "already_running",
                "message": f"Conversion already in progress for {job_id}"
            }

        # Create job with ISBN as ID
        job = ConversionJob(
            job_id=job_id,
            input_file=str(input_path),
            output_dir=str(out_dir)
        )

        # Store in active jobs cache and MongoDB
        with self._lock:
            self._active_jobs[job_id] = job
        self._save_job_to_mongodb(job)

        if async_mode:
            # Start conversion in background thread
            thread = threading.Thread(
                target=self._run_conversion,
                args=(job_id, input_path, out_dir, debug),
                daemon=True
            )
            self._job_threads[job_id] = thread
            thread.start()

            return {
                "job_id": job_id,
                "isbn": job_id,
                "status": "started",
                "message": f"Conversion started for {input_path.name}"
            }
        else:
            # Run synchronously
            self._run_conversion(job_id, input_path, out_dir, debug)
            return self.get_job_status(job_id)

    def _run_conversion(
        self,
        job_id: str,
        input_path: Path,
        output_dir: Path,
        debug: bool = False
    ) -> None:
        """
        Run the actual conversion (called in thread or synchronously).
        Updates MongoDB on each status change for persistence.
        """
        job = self.get_job(job_id)
        if not job:
            logger.error(f"Job not found: {job_id}")
            return

        try:
            # Update job status to RUNNING
            job.status = JobStatus.RUNNING
            job.started_at = datetime.now().isoformat()
            job.message = "Starting conversion..."

            # Save to MongoDB and update active cache
            with self._lock:
                self._active_jobs[job_id] = job
            self._save_job_to_mongodb(job)
            self._notify_progress(job_id, 0, "Starting conversion...")

            # Import here to avoid circular imports
            from epub_pipeline import convert_epub

            # Run conversion with thread-safe context
            # This ensures each concurrent conversion has isolated ID Authority and Reference Mapper
            if THREAD_SAFE_CONTEXT_AVAILABLE:
                with thread_safe_conversion(conversion_id=job_id):
                    result_path = convert_epub(
                        input_path=input_path,
                        out_dir=output_dir,
                        debug=debug,
                        interactive=False  # Disable interactive mode for API
                    )
            else:
                # Fall back to non-thread-safe conversion
                result_path = convert_epub(
                    input_path=input_path,
                    out_dir=output_dir,
                    debug=debug,
                    interactive=False  # Disable interactive mode for API
                )

            # Update job with success
            job.status = JobStatus.COMPLETED
            job.completed_at = datetime.now().isoformat()
            job.progress = 100
            job.message = "Conversion completed successfully"
            job.result = {
                "output_file": str(result_path),
                "output_path": str(result_path),  # Alias for UI compatibility
                "output_dir": str(output_dir),
                "file_size_mb": result_path.stat().st_size / (1024 * 1024) if result_path.exists() else 0,
                "isbn": job_id
            }

            # Upload output files to storage backend
            storage_files = self._upload_to_storage(job_id, result_path, output_dir)
            if storage_files:
                job.result["storage_files"] = storage_files
                job.message = f"Conversion completed - {len(storage_files)} files stored"

            # Save final status to MongoDB
            self._save_job_to_mongodb(job)
            self._notify_progress(job_id, 100, "Conversion completed successfully")

            # Remove from active jobs (completed jobs are in MongoDB)
            with self._lock:
                self._active_jobs.pop(job_id, None)

        except Exception as e:
            # Update job with failure
            job.status = JobStatus.FAILED
            job.completed_at = datetime.now().isoformat()
            job.error = str(e)
            job.message = f"Conversion failed: {str(e)}"

            # Save failure status to MongoDB
            self._save_job_to_mongodb(job)
            self._notify_progress(job_id, -1, f"Conversion failed: {str(e)}")
            logger.exception(f"Conversion failed for job {job_id}")

            # Remove from active jobs
            with self._lock:
                self._active_jobs.pop(job_id, None)

    def _upload_to_storage(
        self,
        isbn: str,
        output_path: Path,
        output_dir: Path
    ) -> List[Dict[str, Any]]:
        """
        Upload conversion outputs to storage backend.

        Args:
            isbn: ISBN identifier for the conversion
            output_path: Path to the main output file (ZIP)
            output_dir: Output directory to check for additional files

        Returns:
            List of uploaded file info dicts
        """
        if not STORAGE_AVAILABLE:
            logger.debug("Storage not available - skipping upload")
            return []

        uploaded = []

        try:
            storage = get_storage()

            if not storage.is_connected():
                logger.warning("Storage not connected - skipping upload")
                return []

            # Upload main output ZIP
            if output_path.exists():
                with open(output_path, 'rb') as f:
                    stored = storage.save_file(
                        isbn=isbn,
                        filename=output_path.name,
                        data=f,
                        content_type="application/zip",
                        metadata={
                            "type": "output_package",
                            "original_path": str(output_path)
                        }
                    )
                    uploaded.append({
                        "filename": stored.filename,
                        "size": stored.size,
                        "type": "output_package"
                    })
                    logger.info(f"Uploaded {output_path.name} to storage ({stored.size} bytes)")

            # Upload validation report if exists
            report_patterns = [
                f"{isbn}_validation_report.xlsx",
                f"{isbn}_report.xlsx"
            ]
            for pattern in report_patterns:
                report_path = output_dir / pattern
                if report_path.exists():
                    with open(report_path, 'rb') as f:
                        stored = storage.save_file(
                            isbn=isbn,
                            filename=report_path.name,
                            data=f,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            metadata={
                                "type": "validation_report",
                                "original_path": str(report_path)
                            }
                        )
                        uploaded.append({
                            "filename": stored.filename,
                            "size": stored.size,
                            "type": "validation_report"
                        })
                        logger.info(f"Uploaded {report_path.name} to storage ({stored.size} bytes)")
                    break  # Only upload first matching report

            # Upload original EPUB (for reference)
            epub_patterns = [f"{isbn}.epub", f"{isbn}.epub3"]
            uploads_dir = output_dir / "uploads"
            for pattern in epub_patterns:
                epub_path = uploads_dir / pattern
                if epub_path.exists():
                    with open(epub_path, 'rb') as f:
                        stored = storage.save_file(
                            isbn=isbn,
                            filename=f"original/{epub_path.name}",
                            data=f,
                            content_type="application/epub+zip",
                            metadata={
                                "type": "original_epub",
                                "original_path": str(epub_path)
                            }
                        )
                        uploaded.append({
                            "filename": stored.filename,
                            "size": stored.size,
                            "type": "original_epub"
                        })
                        logger.info(f"Uploaded original EPUB to storage ({stored.size} bytes)")
                    break

        except Exception as e:
            logger.error(f"Failed to upload files to storage: {e}")

        return uploaded

    def get_job_status(self, job_id: str) -> Dict[str, Any]:
        """
        Get the status of a conversion job.

        Args:
            job_id: The job ID (ISBN) returned from start_conversion

        Returns:
            Dict with job status and details
        """
        job = self.get_job(job_id)
        if not job:
            raise KeyError(f"Job not found: {job_id}")

        return job.to_dict()

    def list_jobs(
        self,
        status: Optional[JobStatus] = None,
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """
        List conversion jobs from MongoDB.

        Args:
            status: Filter by status (optional)
            limit: Maximum number of jobs to return

        Returns:
            List of job dictionaries
        """
        # Get jobs from MongoDB
        status_str = status.value if status else None
        mongodb_jobs = self._list_jobs_from_mongodb(status=status_str, limit=limit)

        # Also include active (in-memory) jobs
        with self._lock:
            active_jobs = [j.to_dict() for j in self._active_jobs.values()]

        # Merge and deduplicate (active jobs take precedence)
        active_ids = {j['job_id'] for j in active_jobs}
        all_jobs = active_jobs + [j for j in mongodb_jobs if j.get('job_id') not in active_ids]

        # Filter by status if specified and MongoDB wasn't available
        if status and not mongodb_jobs:
            all_jobs = [j for j in all_jobs if j.get('status') == status.value]

        # Sort by created_at descending
        all_jobs.sort(key=lambda x: x.get('created_at', ''), reverse=True)

        # Limit results
        return all_jobs[:limit]

    def cancel_job(self, job_id: str) -> Dict[str, Any]:
        """
        Cancel a running conversion job.

        Args:
            job_id: The job ID (ISBN) to cancel

        Returns:
            Dict with cancellation status
        """
        job = self.get_job(job_id)
        if not job:
            raise KeyError(f"Job not found: {job_id}")

        if job.status == JobStatus.RUNNING:
            job.status = JobStatus.CANCELLED
            job.completed_at = datetime.now().isoformat()
            job.message = "Job cancelled by user"

            # Update in MongoDB
            self._save_job_to_mongodb(job)

            # Remove from active jobs
            with self._lock:
                self._active_jobs.pop(job_id, None)

            return {"status": "cancelled", "message": "Job marked as cancelled"}

        elif job.status == JobStatus.PENDING:
            job.status = JobStatus.CANCELLED
            job.message = "Job cancelled before starting"

            # Update in MongoDB
            self._save_job_to_mongodb(job)

            # Remove from active jobs
            with self._lock:
                self._active_jobs.pop(job_id, None)

            return {"status": "cancelled", "message": "Job cancelled before starting"}

        else:
            return {"status": "error", "message": f"Cannot cancel job with status {job.status}"}

    def get_dashboard_data(self) -> Dict[str, Any]:
        """
        Get conversion dashboard data from MongoDB.

        Returns:
            Dict with dashboard statistics and recent conversions
        """
        # Get statistics from MongoDB
        all_jobs = self.list_jobs(limit=1000)  # Get all jobs for stats

        total_jobs = len(all_jobs)
        completed_jobs = len([j for j in all_jobs if j.get('status') == JobStatus.COMPLETED.value])
        failed_jobs = len([j for j in all_jobs if j.get('status') == JobStatus.FAILED.value])
        running_jobs = len([j for j in all_jobs if j.get('status') == JobStatus.RUNNING.value])
        pending_jobs = len([j for j in all_jobs if j.get('status') == JobStatus.PENDING.value])

        # Get recent jobs
        recent_jobs = all_jobs[:10]

        # Try to load dashboard from Excel if exists
        dashboard_file = self.output_dir / "conversion_dashboard.xlsx"
        dashboard_records = []
        if dashboard_file.exists():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(dashboard_file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    record = dict(zip(headers, row))
                    dashboard_records.append(record)
                wb.close()
            except Exception as e:
                logger.warning(f"Could not load dashboard file: {e}")

        return {
            "statistics": {
                "total_jobs": total_jobs,
                "completed": completed_jobs,
                "failed": failed_jobs,
                "running": running_jobs,
                "pending": pending_jobs,
                "success_rate": (completed_jobs / total_jobs * 100) if total_jobs > 0 else 0
            },
            "recent_jobs": recent_jobs,
            "dashboard_records": dashboard_records[-20:],  # Last 20 records
            "timestamp": datetime.now().isoformat()
        }

    def get_conversion_result(self, job_id: str) -> Dict[str, Any]:
        """
        Get detailed result of a completed conversion.

        Args:
            job_id: The job ID (ISBN)

        Returns:
            Dict with conversion result details
        """
        job = self.get_job(job_id)
        if not job:
            raise KeyError(f"Job not found: {job_id}")

        if job.status != JobStatus.COMPLETED:
            return {
                "status": "error",
                "message": f"Job is not completed. Current status: {job.status}"
            }

        result = job.result or {}
        output_file = Path(result.get("output_file", ""))

        # Get additional details from the output
        output_path_str = str(output_file) if output_file.exists() else None
        details = {
            "job_id": job_id,
            "input_file": job.input_file,
            "output_file": output_path_str,
            "output_path": output_path_str,  # Alias for UI compatibility
            "output_dir": job.output_dir,
            "started_at": job.started_at,
            "completed_at": job.completed_at,
            "duration_seconds": None,
            "file_size_mb": result.get("file_size_mb", 0),
            "validation_report": None,
            "intermediate_files": []
        }

        # Calculate duration
        if job.started_at and job.completed_at:
            start = datetime.fromisoformat(job.started_at)
            end = datetime.fromisoformat(job.completed_at)
            details["duration_seconds"] = (end - start).total_seconds()

        # Find associated files
        output_dir = Path(job.output_dir)
        isbn = Path(job.input_file).stem

        # Look for validation report
        report_path = output_dir / f"{isbn}_validation_report.xlsx"
        if report_path.exists():
            details["validation_report"] = str(report_path)

        # Look for intermediate files
        intermediate_dir = output_dir / f"{isbn}_intermediate"
        if intermediate_dir.exists():
            details["intermediate_files"] = [
                str(f) for f in intermediate_dir.iterdir() if f.is_file()
            ]

        return details

    def register_progress_callback(
        self,
        job_id: str,
        callback: Callable[[int, str], None]
    ) -> None:
        """
        Register a callback for progress updates.

        Args:
            job_id: The job ID to monitor
            callback: Function to call with (progress, message)
        """
        if job_id not in self._progress_callbacks:
            self._progress_callbacks[job_id] = []
        self._progress_callbacks[job_id].append(callback)

    def _notify_progress(self, job_id: str, progress: int, message: str) -> None:
        """Notify all registered callbacks of progress update."""
        if job_id in self._progress_callbacks:
            for callback in self._progress_callbacks[job_id]:
                try:
                    callback(progress, message)
                except Exception as e:
                    logger.warning(f"Progress callback failed: {e}")

    def cleanup_old_jobs(self, max_age_hours: int = 24) -> int:
        """
        Clean up old completed/failed jobs from active memory cache.
        Note: Jobs in MongoDB are kept for historical records.

        Args:
            max_age_hours: Maximum age of jobs to keep in memory

        Returns:
            Number of jobs cleaned up from memory
        """
        cutoff = datetime.now()
        cleaned = 0

        with self._lock:
            jobs_to_remove = []
            for job_id, job in self._active_jobs.items():
                if job.status in [JobStatus.COMPLETED, JobStatus.FAILED, JobStatus.CANCELLED]:
                    if job.completed_at:
                        completed = datetime.fromisoformat(job.completed_at)
                        age_hours = (cutoff - completed).total_seconds() / 3600
                        if age_hours > max_age_hours:
                            jobs_to_remove.append(job_id)

            for job_id in jobs_to_remove:
                del self._active_jobs[job_id]
                if job_id in self._progress_callbacks:
                    del self._progress_callbacks[job_id]
                cleaned += 1

        logger.info(f"Cleaned up {cleaned} old jobs from memory")
        return cleaned


# Singleton instance for easy access
_api_instance: Optional[ConversionAPI] = None


def get_api(output_dir: Optional[Path] = None) -> ConversionAPI:
    """Get or create the ConversionAPI singleton instance."""
    global _api_instance
    if _api_instance is None:
        _api_instance = ConversionAPI(output_dir)
    return _api_instance
