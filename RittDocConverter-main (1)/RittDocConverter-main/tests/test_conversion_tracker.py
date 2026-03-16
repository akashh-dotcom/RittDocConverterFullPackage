"""
Comprehensive tests for conversion_tracker.py

Tests for:
- ConversionStatus, ConversionType, TemplateType enums
- ConversionMetadata dataclass
- ConversionTracker class
- Excel dashboard generation
- MongoDB integration
"""

import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch

import pytest
import openpyxl

from conversion_tracker import (
    ConversionMetadata,
    ConversionStatus,
    ConversionTracker,
    ConversionType,
    TemplateType,
    track_conversion,
)


class TestConversionStatus:
    """Tests for ConversionStatus enum."""

    def test_status_values(self):
        """Test all status enum values."""
        assert ConversionStatus.IN_PROGRESS.value == "In Progress"
        assert ConversionStatus.SUCCESS.value == "Success"
        assert ConversionStatus.FAILURE.value == "Failure"
        assert ConversionStatus.PARTIAL.value == "Partial Success"


class TestConversionType:
    """Tests for ConversionType enum."""

    def test_type_values(self):
        """Test all conversion type values."""
        assert ConversionType.PDF.value == "PDF"
        assert ConversionType.EPUB.value == "ePub"
        assert ConversionType.DOCX.value == "DOCX"


class TestTemplateType:
    """Tests for TemplateType enum."""

    def test_template_values(self):
        """Test all template type values."""
        assert TemplateType.SINGLE_COLUMN.value == "Single Column"
        assert TemplateType.DOUBLE_COLUMN.value == "Double Column"
        assert TemplateType.MIXED.value == "Mixed"
        assert TemplateType.UNKNOWN.value == "Unknown"


class TestConversionMetadata:
    """Tests for ConversionMetadata dataclass."""

    def test_metadata_creation_minimal(self):
        """Test creating metadata with minimal fields."""
        metadata = ConversionMetadata(filename="test.epub")
        assert metadata.filename == "test.epub"
        assert metadata.status == ConversionStatus.IN_PROGRESS
        assert metadata.progress_percent == 0

    def test_metadata_creation_full(self):
        """Test creating metadata with all fields."""
        start = datetime.now()
        metadata = ConversionMetadata(
            filename="test.epub",
            isbn="9781234567890",
            publisher="Test Publisher",
            start_time=start,
            status=ConversionStatus.SUCCESS,
            conversion_type=ConversionType.EPUB,
            num_chapters=10,
            num_pages=200,
            num_vector_images=5,
            num_raster_images=15,
            num_tables=8,
            title="Test Book",
            authors=["Author 1", "Author 2"],
        )

        assert metadata.filename == "test.epub"
        assert metadata.isbn == "9781234567890"
        assert metadata.publisher == "Test Publisher"
        assert metadata.status == ConversionStatus.SUCCESS
        assert metadata.num_chapters == 10
        assert len(metadata.authors) == 2

    def test_duration_seconds_none(self):
        """Test duration when end_time not set."""
        metadata = ConversionMetadata(filename="test.epub")
        assert metadata.duration_seconds() is None

    def test_duration_seconds_calculated(self):
        """Test duration calculation."""
        from datetime import timedelta

        start = datetime.now()
        end = start + timedelta(seconds=120)

        metadata = ConversionMetadata(
            filename="test.epub",
            start_time=start,
            end_time=end
        )

        duration = metadata.duration_seconds()
        assert duration == 120

    def test_to_row(self):
        """Test converting metadata to Excel row."""
        metadata = ConversionMetadata(
            filename="test.epub",
            isbn="9781234567890",
            title="Test Book",
            publisher="Test Publisher",
            authors=["Author 1", "Author 2"],
            num_chapters=5,
            num_vector_images=10,
            num_raster_images=20,
            num_tables=3,
        )

        row = metadata.to_row()
        assert isinstance(row, list)
        assert row[0] == "test.epub"
        assert row[1] == "9781234567890"
        assert row[2] == "Test Book"
        assert row[3] == "Test Publisher"
        assert "Author 1" in row[4]
        assert row[16] == 30  # Total images

    def test_to_row_with_completed_conversion(self):
        """Test row format for completed conversion."""
        from datetime import timedelta

        start = datetime.now()
        end = start + timedelta(seconds=60)

        metadata = ConversionMetadata(
            filename="test.epub",
            start_time=start,
            end_time=end,
            status=ConversionStatus.SUCCESS
        )

        row = metadata.to_row()
        assert row[7] == "60s"  # Duration
        assert row[8] == "Success"  # Status

    def test_to_dict(self):
        """Test converting metadata to dictionary."""
        metadata = ConversionMetadata(
            filename="test.epub",
            isbn="9781234567890",
            title="Test Book",
            num_chapters=5,
            num_vector_images=10,
            num_raster_images=15
        )

        data = metadata.to_dict()
        assert isinstance(data, dict)
        assert data['filename'] == "test.epub"
        assert data['isbn'] == "9781234567890"
        assert data['title'] == "Test Book"
        assert data['total_images'] == 25
        assert 'start_time' in data

    def test_authors_list_in_row(self):
        """Test authors are formatted correctly in row."""
        metadata = ConversionMetadata(
            filename="test.epub",
            authors=["John Doe", "Jane Smith"]
        )

        row = metadata.to_row()
        assert "John Doe, Jane Smith" in row[4]

    def test_empty_authors_in_row(self):
        """Test empty authors list in row."""
        metadata = ConversionMetadata(filename="test.epub")
        row = metadata.to_row()
        assert row[4] == ""


class TestConversionTracker:
    """Tests for ConversionTracker class."""

    @pytest.fixture
    def temp_output_dir(self, tmp_path):
        """Create temporary output directory."""
        output_dir = tmp_path / "output"
        output_dir.mkdir()
        return output_dir

    @pytest.fixture
    def tracker(self, temp_output_dir):
        """Create ConversionTracker instance."""
        return ConversionTracker(temp_output_dir)

    def test_tracker_initialization(self, temp_output_dir):
        """Test tracker initialization."""
        tracker = ConversionTracker(temp_output_dir)
        assert tracker.output_dir == temp_output_dir
        assert tracker.excel_path == temp_output_dir / "conversion_dashboard.xlsx"
        assert tracker.current_metadata is None

    def test_start_conversion(self, tracker):
        """Test starting a conversion."""
        metadata = tracker.start_conversion(
            filename="test.epub",
            conversion_type=ConversionType.EPUB,
            isbn="9781234567890"
        )

        assert metadata is not None
        assert metadata.filename == "test.epub"
        assert metadata.conversion_type == ConversionType.EPUB
        assert metadata.isbn == "9781234567890"
        assert tracker.current_metadata == metadata

    def test_update_progress(self, tracker):
        """Test updating progress."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.update_progress(50)

        assert tracker.current_metadata.progress_percent == 50

    def test_update_progress_with_status(self, tracker):
        """Test updating progress with status change."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.update_progress(75, status=ConversionStatus.SUCCESS)

        assert tracker.current_metadata.progress_percent == 75
        assert tracker.current_metadata.status == ConversionStatus.SUCCESS

    def test_update_progress_clamps_to_100(self, tracker):
        """Test progress is clamped to 100."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.update_progress(150)

        assert tracker.current_metadata.progress_percent == 100

    def test_update_progress_clamps_to_0(self, tracker):
        """Test progress is clamped to 0."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.update_progress(-10)

        assert tracker.current_metadata.progress_percent == 0

    def test_complete_conversion_success(self, tracker):
        """Test completing a successful conversion."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(
            status=ConversionStatus.SUCCESS,
            output_path="/path/to/output.zip",
            output_size_mb=5.5
        )

        assert tracker.current_metadata.status == ConversionStatus.SUCCESS
        assert tracker.current_metadata.progress_percent == 100
        assert tracker.current_metadata.end_time is not None
        assert tracker.current_metadata.output_path == "/path/to/output.zip"
        assert tracker.current_metadata.output_size_mb == 5.5

    def test_complete_conversion_failure(self, tracker):
        """Test completing a failed conversion."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(
            status=ConversionStatus.FAILURE,
            error_message="Test error message"
        )

        assert tracker.current_metadata.status == ConversionStatus.FAILURE
        assert tracker.current_metadata.error_message == "Test error message"
        assert tracker.current_metadata.end_time is not None

    def test_complete_conversion_no_active(self, tracker):
        """Test completing conversion with no active conversion."""
        # Should not raise, just log warning
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

    def test_save_to_excel_creates_file(self, tracker):
        """Test that Excel file is created."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        assert tracker.excel_path.exists()

    def test_save_to_excel_headers(self, tracker):
        """Test Excel file has correct headers."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        assert "Filename" in headers
        assert "ISBN" in headers
        assert "Status" in headers
        assert "# Tables" in headers
        wb.close()

    def test_save_to_excel_appends_data(self, tracker):
        """Test that data is appended to existing file."""
        # First conversion
        tracker.start_conversion("test1.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        # Second conversion
        tracker.start_conversion("test2.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        # Should have header + 2 data rows
        assert ws.max_row >= 3
        wb.close()

    def test_save_to_excel_updates_existing(self, tracker):
        """Test that existing row is updated on progress updates."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.update_progress(50)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active
        initial_rows = ws.max_row
        wb.close()

        tracker.update_progress(75)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active
        # Should not add new row, just update existing
        assert ws.max_row == initial_rows
        wb.close()

    def test_format_header_row(self, tracker):
        """Test header row formatting."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        # Check first cell has header formatting
        first_cell = ws.cell(1, 1)
        assert first_cell.font.bold is True
        # RGB color may have different prefix (FF or 00)
        assert first_cell.fill.start_color.rgb in ["FF366092", "00366092"]
        wb.close()

    def test_format_data_row_success(self, tracker):
        """Test data row formatting for success status."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        # Status column should have success color (may have FF or 00 prefix)
        status_cell = ws.cell(2, 9)
        assert status_cell.fill.start_color.rgb in ["FFC6EFCE", "00C6EFCE"]
        wb.close()

    def test_format_data_row_failure(self, tracker):
        """Test data row formatting for failure status."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.FAILURE)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        # Status column should have failure color (may have FF or 00 prefix)
        status_cell = ws.cell(2, 9)
        assert status_cell.fill.start_color.rgb in ["FFFFC7CE", "00FFC7CE"]
        wb.close()

    def test_format_data_row_in_progress(self, tracker):
        """Test data row formatting for in-progress status."""
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.update_progress(50)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        # Status column should have in-progress color (may have FF or 00 prefix)
        status_cell = ws.cell(2, 9)
        assert status_cell.fill.start_color.rgb in ["FFFFEB9C", "00FFEB9C"]
        wb.close()

    def test_push_to_mongodb_not_available(self, tracker):
        """Test MongoDB push when not available."""
        with patch('conversion_tracker.MONGODB_AVAILABLE', False):
            tracker.start_conversion("test.epub", ConversionType.EPUB)
            # Should not raise, just skip
            tracker.complete_conversion(status=ConversionStatus.SUCCESS)

    def test_push_to_mongodb_available(self, tracker):
        """Test MongoDB push when available."""
        with patch('conversion_tracker.MONGODB_AVAILABLE', True):
            with patch('conversion_tracker.get_mongodb_client') as mock_client:
                mock_client.return_value.push_conversion.return_value = True

                tracker.start_conversion("test.epub", ConversionType.EPUB)
                tracker.complete_conversion(status=ConversionStatus.SUCCESS)

                # Should have called push_conversion
                assert mock_client.return_value.push_conversion.called

    def test_push_to_mongodb_failure_non_fatal(self, tracker):
        """Test MongoDB push failure doesn't stop conversion."""
        with patch('conversion_tracker.MONGODB_AVAILABLE', True):
            with patch('conversion_tracker.get_mongodb_client') as mock_client:
                mock_client.return_value.push_conversion.side_effect = Exception("DB error")

                tracker.start_conversion("test.epub", ConversionType.EPUB)
                # Should not raise
                tracker.complete_conversion(status=ConversionStatus.SUCCESS)

    def test_get_statistics_no_file(self, tracker):
        """Test getting statistics when Excel file doesn't exist."""
        stats = tracker.get_statistics()
        assert stats == {}

    def test_get_statistics_with_data(self, tracker):
        """Test getting statistics from Excel file."""
        # Create some conversions
        tracker.start_conversion("test1.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        tracker.start_conversion("test2.epub", ConversionType.PDF)
        tracker.complete_conversion(status=ConversionStatus.FAILURE)

        stats = tracker.get_statistics()
        assert stats['total_conversions'] >= 2
        assert stats['successful'] >= 1
        assert stats['failed'] >= 1
        assert stats['epub_conversions'] >= 1

    def test_auto_size_columns(self, tracker):
        """Test column auto-sizing."""
        tracker.start_conversion("test.epub", ConversionType.EPUB, isbn="9781234567890")
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        # Columns should have adjusted widths
        for col in ws.column_dimensions.values():
            assert col.width > 0
            assert col.width <= 50  # Capped at 50

        wb.close()

    def test_tracker_with_all_metadata_fields(self, tracker):
        """Test tracker with all possible metadata fields."""
        metadata = tracker.start_conversion(
            filename="comprehensive_test.epub",
            conversion_type=ConversionType.EPUB,
            isbn="9781234567890",
            publisher="Test Publisher",
            template_type=TemplateType.DOUBLE_COLUMN,
            num_chapters=15,
            num_pages=300,
            num_vector_images=25,
            num_raster_images=50,
            num_tables=12,
            num_equations=8,
            title="Comprehensive Test Book",
            authors=["Author One", "Author Two", "Author Three"],
            language="en-US"
        )

        tracker.complete_conversion(
            status=ConversionStatus.SUCCESS,
            output_path="/output/test.zip",
            output_size_mb=10.5
        )

        # Verify all fields are saved
        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active

        row_data = [cell.value for cell in ws[2]]
        assert "comprehensive_test.epub" in row_data
        assert "9781234567890" in row_data
        assert "Test Publisher" in row_data
        wb.close()


class TestTrackConversionFunction:
    """Tests for track_conversion convenience function."""

    def test_track_conversion_creates_tracker(self, tmp_path):
        """Test that track_conversion creates and starts a tracker."""
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        tracker = track_conversion(
            output_dir=output_dir,
            filename="test.epub",
            conversion_type=ConversionType.EPUB,
            isbn="9781234567890"
        )

        assert isinstance(tracker, ConversionTracker)
        assert tracker.current_metadata is not None
        assert tracker.current_metadata.filename == "test.epub"
        assert tracker.current_metadata.isbn == "9781234567890"

    def test_track_conversion_with_additional_metadata(self, tmp_path):
        """Test track_conversion with extra metadata."""
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        tracker = track_conversion(
            output_dir=output_dir,
            filename="test.epub",
            conversion_type=ConversionType.EPUB,
            publisher="Test Publisher",
            num_chapters=10
        )

        assert tracker.current_metadata.publisher == "Test Publisher"
        assert tracker.current_metadata.num_chapters == 10


class TestExcelWorkbookHandling:
    """Tests for Excel workbook handling edge cases."""

    @pytest.fixture
    def temp_output_dir(self, tmp_path):
        """Create temporary output directory."""
        output_dir = tmp_path / "output"
        output_dir.mkdir()
        return output_dir

    def test_creates_new_workbook_if_not_exists(self, temp_output_dir):
        """Test creating new workbook."""
        tracker = ConversionTracker(temp_output_dir)
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        assert tracker.excel_path.exists()

    def test_loads_existing_workbook(self, temp_output_dir):
        """Test loading existing workbook."""
        tracker = ConversionTracker(temp_output_dir)

        # Create initial workbook
        tracker.start_conversion("test1.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        # Create new tracker with same output dir
        tracker2 = ConversionTracker(temp_output_dir)
        tracker2.start_conversion("test2.epub", ConversionType.EPUB)
        tracker2.complete_conversion(status=ConversionStatus.SUCCESS)

        # Should have both entries
        wb = openpyxl.load_workbook(tracker.excel_path)
        ws = wb.active
        assert ws.max_row >= 3  # Header + 2 rows
        wb.close()

    def test_handles_missing_sheet_name(self, temp_output_dir):
        """Test handling workbook without expected sheet."""
        # Create workbook with different sheet name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DifferentSheet"
        excel_path = temp_output_dir / "conversion_dashboard.xlsx"
        wb.save(excel_path)
        wb.close()

        # Tracker should create new sheet
        tracker = ConversionTracker(temp_output_dir)
        tracker.start_conversion("test.epub", ConversionType.EPUB)
        tracker.complete_conversion(status=ConversionStatus.SUCCESS)

        wb = openpyxl.load_workbook(excel_path)
        assert "Conversions" in wb.sheetnames
        wb.close()


class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_metadata_with_none_values(self):
        """Test metadata with None values."""
        metadata = ConversionMetadata(
            filename="test.epub",
            isbn=None,
            publisher=None,
            title=None
        )

        row = metadata.to_row()
        assert row[1] == ""  # ISBN
        assert row[2] == ""  # Title
        assert row[3] == ""  # Publisher

    def test_metadata_with_special_characters(self):
        """Test metadata with special characters."""
        metadata = ConversionMetadata(
            filename="test-file@2023!.epub",
            title="Book: A Story of Success & Failure",
            publisher="Publisher's House"
        )

        row = metadata.to_row()
        # Should handle special characters
        assert "test-file@2023!" in row[0]
        assert "Success & Failure" in row[2]

    def test_very_long_filename(self):
        """Test metadata with very long filename."""
        long_filename = "a" * 300 + ".epub"
        metadata = ConversionMetadata(filename=long_filename)

        row = metadata.to_row()
        # Should not crash
        assert len(row) > 0