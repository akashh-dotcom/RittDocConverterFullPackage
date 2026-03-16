"""
DTD Compliance Checker and Auto-Fixer for RittDoc XML

ARCHITECTURE ROLE:
    This module is used DURING CONVERSION (Phase 3) on in-memory XML.
    It operates on the book element before writing to disk.

    For POST-PROCESSING on individual chapter files, see:
    comprehensive_dtd_fixer.py

USAGE:
    from dtd_compliance import validate_and_fix_dtd_compliance
    report = validate_and_fix_dtd_compliance(book_elem)

This module validates XML against RittDoc/DocBook DTD requirements and
automatically fixes violations where possible.

Based on comprehensive analysis of:
- RITTDOCdtd/v1.1/dbpoolx.mod (pool elements)
- RITTDOCdtd/v1.1/dbhierx.mod (hierarchical elements)
- RITTDOCdtd/v1.1/calstblx.dtd (table elements)
"""

import logging
import re
from typing import Dict, List, Set, Tuple, Optional, Any
from lxml import etree
from dataclasses import dataclass, field

from id_authority import next_available_sect1_id

logger = logging.getLogger(__name__)


@dataclass
class ValidationIssue:
    """Represents a DTD validation issue"""
    element_tag: str
    element_id: Optional[str]
    issue_type: str
    description: str
    severity: str  # 'error', 'warning'
    auto_fixed: bool = False
    fix_description: Optional[str] = None


@dataclass
class ComplianceReport:
    """Report of all validation issues found"""
    issues: List[ValidationIssue] = field(default_factory=list)
    fixed_count: int = 0
    unfixed_count: int = 0

    def add_issue(self, issue: ValidationIssue):
        self.issues.append(issue)
        if issue.auto_fixed:
            self.fixed_count += 1
        else:
            self.unfixed_count += 1

    def get_summary(self) -> str:
        return (f"DTD Compliance: {len(self.issues)} issues found, "
                f"{self.fixed_count} auto-fixed, {self.unfixed_count} require attention")


# =============================================================================
# DTD CONTENT MODEL DEFINITIONS
# =============================================================================

# Elements that MUST NOT be empty (require at least one valid child)
ELEMENTS_REQUIRING_CONTENT = {
    # Hierarchical elements (require title + content from divcomponent.mix)
    'sect1', 'sect2', 'sect3', 'sect4', 'sect5', 'section', 'simplesect',
    'chapter', 'appendix', 'preface', 'part', 'dedication', 'colophon',

    # Lists (require at least one item)
    'itemizedlist', 'orderedlist', 'variablelist', 'simplelist',
    'segmentedlist', 'calloutlist', 'glosslist',

    # List items and entries (require content from component.mix)
    'listitem', 'varlistentry', 'callout', 'glossentry',

    # Admonitions (require content from admon.mix)
    'note', 'warning', 'caution', 'important', 'tip',

    # Block elements
    'blockquote', 'epigraph', 'footnote', 'sidebar',
    'abstract', 'authorblurb', 'personblurb', 'legalnotice',

    # Figures and examples (require content)
    'figure', 'informalfigure', 'example', 'informalexample',

    # Media objects
    'mediaobject', 'inlinemediaobject', 'imageobject', 'videoobject', 'audioobject',

    # Tables (require proper structure)
    'table', 'informaltable', 'tgroup', 'thead', 'tfoot', 'tbody', 'row',

    # Bibliography/Glossary
    'bibliography', 'bibliodiv', 'glossary', 'glossdiv', 'glossdef',

    # Index
    'indexdiv', 'indexentry',

    # TOC elements
    'tocpart', 'tocchap', 'toclevel1', 'toclevel2', 'toclevel3', 'toclevel4', 'toclevel5',

    # Reference entries
    'refentry', 'refnamediv', 'refsynopsisdiv', 'refsection', 'refsect1', 'refsect2', 'refsect3',

    # Procedures
    'procedure', 'step', 'substeps', 'stepalternatives',

    # Q&A
    'question', 'qandaentry',

    # Messages
    'msgset', 'msgentry', 'simplemsgentry', 'msg', 'msgmain', 'msgexplan',

    # Synopses
    'cmdsynopsis', 'funcsynopsis', 'classsynopsis',

    # Other structural
    'highlights', 'revhistory', 'authorgroup', 'copyright',
}

# Elements that CAN be empty (have * or ? in content model)
ELEMENTS_CAN_BE_EMPTY = {
    'para', 'simpara', 'title', 'titleabbrev', 'subtitle',
    'programlisting', 'literallayout', 'screen', 'synopsis',
    'caption', 'remark', 'address', 'entry',  # table entry can be empty
    'funcsynopsisinfo', 'classsynopsisinfo',
    'phrase', 'emphasis', 'literal', 'code', 'command',
    'computeroutput', 'userinput', 'filename', 'option',
}

# Elements with EMPTY content model (no children allowed)
EMPTY_ELEMENTS = {
    'colspec', 'spanspec', 'area', 'co', 'coref',
    'graphic', 'inlinegraphic', 'sbr', 'void', 'varargs',
    'footnoteref', 'xref', 'anchor', 'beginpage',
    'videodata', 'audiodata', 'imagedata', 'textdata',
}

# Element ordering requirements (element -> ordered list of allowed children)
# Format: element -> [(child_tag, required, position_constraint)]
# position_constraint: 'first', 'early', 'before:X', 'after:X', None
ELEMENT_ORDERING = {
    # Elements where beginpage comes BEFORE title per DTD:
    'chapter': [
        ('beginpage', False, 'first'),
        ('chapterinfo', False, 'early'),
        ('title', True, 'after:beginpage'),
    ],
    'appendix': [
        ('beginpage', False, 'first'),
        ('appendixinfo', False, 'early'),
        ('title', True, 'after:beginpage'),
    ],
    'preface': [
        ('beginpage', False, 'first'),
        ('prefaceinfo', False, 'early'),
        ('title', True, 'after:beginpage'),
    ],
    'part': [
        ('beginpage', False, 'first'),
        ('partinfo', False, 'early'),
        ('title', True, 'after:beginpage'),
    ],
    'subpart': [
        ('beginpage', False, 'first'),
        ('partinfo', False, 'early'),
        ('title', True, 'after:beginpage'),
    ],
    'toc': [
        ('beginpage', False, 'first'),
        ('title', False, 'after:beginpage'),
    ],
    'lot': [
        ('beginpage', False, 'first'),
        ('title', False, 'after:beginpage'),
    ],
    # Elements without beginpage per DTD:
    'dedication': [
        ('risinfo', False, 'first'),
        ('title', False, 'after:risinfo'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'colophon': [
        # No beginpage per DTD: (sect.title.content?, textobject.mix+)
        ('title', False, 'first'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    # Sections: beginpage is in divcomponent.mix AFTER title, not before
    # DTD: (sectNinfo?, title, subtitle?, titleabbrev?, divcomponent.mix+)
    'sect1': [
        ('sect1info', False, 'first'),
        ('title', True, 'early'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'sect2': [
        ('sect2info', False, 'first'),
        ('title', True, 'early'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'sect3': [
        ('sect3info', False, 'first'),
        ('title', True, 'early'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'sect4': [
        ('sect4info', False, 'first'),
        ('title', True, 'early'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'sect5': [
        ('sect5info', False, 'first'),
        ('title', True, 'early'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'section': [
        ('sectioninfo', False, 'first'),
        ('title', True, 'early'),
        ('subtitle', False, 'after:title'),
        ('titleabbrev', False, 'after:subtitle'),
    ],
    'figure': [
        ('blockinfo', False, 'first'),
        ('title', True, 'early'),
    ],
    'table': [
        ('blockinfo', False, 'first'),
        ('title', False, 'early'),
        # Note: tgroup is required only when NOT using mediaobject/graphic content model.
        # Tables can use (tgroup+) | (mediaobject+) | (graphic+).
        # The check_table_structure() function handles this distinction.
        # Marking tgroup as optional here to avoid false positives for mediaobject tables.
        ('tgroup', False, None),
    ],
    'tgroup': [
        ('colspec', False, 'first'),
        ('spanspec', False, 'after:colspec'),
        ('thead', False, 'before:tbody'),
        ('tfoot', False, 'before:tbody'),
        ('tbody', True, None),
    ],
}

# Valid content elements for sections (divcomponent.mix)
DIVCOMPONENT_MIX = {
    'para', 'simpara', 'formalpara', 'programlisting', 'literallayout',
    'screen', 'synopsis', 'address', 'blockquote', 'epigraph',
    'figure', 'informalfigure', 'table', 'informaltable',
    'example', 'informalexample', 'equation', 'informalequation',
    'procedure', 'sidebar', 'mediaobject', 'graphic',
    'itemizedlist', 'orderedlist', 'variablelist', 'simplelist',
    'segmentedlist', 'calloutlist', 'glosslist',
    'note', 'warning', 'caution', 'important', 'tip',
    'bridgehead', 'remark', 'highlights', 'abstract',
    'cmdsynopsis', 'funcsynopsis', 'classsynopsis',
    'revhistory', 'task', 'productionset', 'constraintdef',
    'msgset', 'qandaset', 'anchor', 'indexterm', 'beginpage',
}

# Elements that don't count as "valid content" for sections
NON_CONTENT_ELEMENTS = {
    'title', 'titleabbrev', 'subtitle', 'anchor', 'indexterm', 'beginpage',
    'sect1info', 'sect2info', 'sect3info', 'sect4info', 'sect5info',
    'sectioninfo', 'chapterinfo', 'appendixinfo', 'prefaceinfo',
    'blockinfo', 'objectinfo',
}

# Valid content for list items (component.mix)
COMPONENT_MIX = {
    'para', 'simpara', 'formalpara', 'programlisting', 'literallayout',
    'screen', 'synopsis', 'address', 'blockquote', 'epigraph',
    'figure', 'informalfigure', 'table', 'informaltable',
    'example', 'informalexample', 'equation', 'informalequation',
    'procedure', 'sidebar', 'mediaobject', 'graphic',
    'itemizedlist', 'orderedlist', 'variablelist', 'simplelist',
    'segmentedlist', 'calloutlist', 'glosslist',
    'note', 'warning', 'caution', 'important', 'tip',
    'bridgehead', 'remark', 'highlights', 'abstract',
    'cmdsynopsis', 'funcsynopsis', 'classsynopsis',
    'revhistory', 'task', 'productionset', 'constraintdef',
    'msgset', 'qandaset', 'anchor', 'indexterm', 'beginpage',
}

# Valid content for admonitions (admon.mix)
ADMON_MIX = {
    'para', 'simpara', 'formalpara', 'programlisting', 'literallayout',
    'screen', 'synopsis', 'address', 'blockquote', 'epigraph',
    'figure', 'informalfigure', 'table', 'informaltable',
    'example', 'informalexample', 'equation', 'informalequation',
    'procedure', 'sidebar', 'mediaobject', 'graphic',
    'itemizedlist', 'orderedlist', 'variablelist', 'simplelist',
    'segmentedlist', 'calloutlist', 'glosslist',
    'bridgehead', 'remark', 'highlights', 'abstract',
    'cmdsynopsis', 'funcsynopsis', 'classsynopsis',
    'revhistory', 'task', 'productionset', 'constraintdef',
    'anchor', 'indexterm', 'beginpage',
}

# Valid content for dedication and legalnotice (legalnotice.mix)
# Per DTD: (%list.class; | %admon.class; | %linespecific.class; | %para.class; | blockquote | %ndxterm.class; | beginpage)
# Note: anchor is NOT allowed in legalnotice.mix
LEGALNOTICE_MIX = {
    # list.class
    'glosslist', 'itemizedlist', 'orderedlist',
    # admon.class
    'caution', 'important', 'note', 'tip', 'warning',
    # linespecific.class
    'literallayout', 'programlisting', 'screen', 'synopsis', 'address',
    # para.class
    'formalpara', 'para', 'simpara',
    # blockquote
    'blockquote',
    # ndxterm.class
    'indexterm',
    # beginpage
    'beginpage',
}


# =============================================================================
# VALIDATION FUNCTIONS
# =============================================================================

def check_empty_element(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Check if an element that requires content is empty.
    Returns True if element is valid (has content), False if empty.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    # Skip elements that can be empty
    if tag in ELEMENTS_CAN_BE_EMPTY or tag in EMPTY_ELEMENTS:
        return True

    # Check if element requires content
    if tag not in ELEMENTS_REQUIRING_CONTENT:
        return True

    # Count valid children (excluding metadata/anchors)
    valid_children = [child for child in elem if child.tag not in NON_CONTENT_ELEMENTS]

    if not valid_children:
        # Element is empty - try to auto-fix
        fixed = fix_empty_element(elem)
        issue = ValidationIssue(
            element_tag=tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"Element <{tag}> is empty but requires content",
            severity='error',
            auto_fixed=fixed,
            fix_description="Added empty <para/>" if fixed else None
        )
        report.add_issue(issue)
        return fixed

    return True


def fix_empty_element(elem: etree.Element) -> bool:
    """
    Fix an empty element by adding appropriate default content.
    Returns True if fixed successfully.
    """
    tag = elem.tag

    # Sections and chapters: add empty para after title
    if tag in {'sect1', 'sect2', 'sect3', 'sect4', 'sect5', 'section', 'simplesect',
               'chapter', 'appendix', 'preface', 'dedication', 'colophon'}:
        # Find insertion point after title and metadata
        insert_pos = 0
        for i, child in enumerate(elem):
            if child.tag in NON_CONTENT_ELEMENTS:
                insert_pos = i + 1
            else:
                break
        para = etree.Element('para')
        elem.insert(insert_pos, para)
        logger.debug(f"Added empty para to {tag} id={elem.get('id', 'unknown')}")
        return True

    # List items: add empty para
    if tag == 'listitem':
        para = etree.Element('para')
        elem.append(para)
        logger.debug(f"Added empty para to listitem")
        return True

    # Lists: add empty listitem with para
    if tag in {'itemizedlist', 'orderedlist'}:
        listitem = etree.SubElement(elem, 'listitem')
        para = etree.SubElement(listitem, 'para')
        logger.debug(f"Added empty listitem to {tag}")
        return True

    # Variable list: add varlistentry with term and listitem
    if tag == 'variablelist':
        varlistentry = etree.SubElement(elem, 'varlistentry')
        term = etree.SubElement(varlistentry, 'term')
        listitem = etree.SubElement(varlistentry, 'listitem')
        para = etree.SubElement(listitem, 'para')
        logger.debug(f"Added empty varlistentry to variablelist")
        return True

    # Simple list: add empty member
    if tag == 'simplelist':
        member = etree.SubElement(elem, 'member')
        logger.debug(f"Added empty member to simplelist")
        return True

    # Admonitions: add empty para
    if tag in {'note', 'warning', 'caution', 'important', 'tip'}:
        para = etree.Element('para')
        elem.append(para)
        logger.debug(f"Added empty para to {tag}")
        return True

    # Blockquote/epigraph: add empty para
    if tag in {'blockquote', 'epigraph', 'abstract', 'legalnotice',
               'authorblurb', 'personblurb', 'highlights'}:
        para = etree.Element('para')
        elem.append(para)
        logger.debug(f"Added empty para to {tag}")
        return True

    # Footnote: add empty para
    if tag == 'footnote':
        para = etree.Element('para')
        elem.append(para)
        logger.debug(f"Added empty para to footnote")
        return True

    # Sidebar: add empty para
    if tag == 'sidebar':
        para = etree.Element('para')
        elem.append(para)
        logger.debug(f"Added empty para to sidebar")
        return True

    # Glossary entries
    if tag == 'glossentry':
        if elem.find('glossterm') is None:
            glossterm = etree.SubElement(elem, 'glossterm')
        if elem.find('glossdef') is None:
            glossdef = etree.SubElement(elem, 'glossdef')
            para = etree.SubElement(glossdef, 'para')
        return True

    if tag == 'glossdef':
        para = etree.Element('para')
        elem.append(para)
        return True

    # Callout: add empty para
    if tag == 'callout':
        para = etree.Element('para')
        elem.append(para)
        return True

    # Q&A
    if tag == 'question':
        para = etree.Element('para')
        elem.append(para)
        return True

    # Procedure: add empty step
    if tag == 'procedure':
        step = etree.SubElement(elem, 'step')
        para = etree.SubElement(step, 'para')
        return True

    if tag == 'step':
        para = etree.Element('para')
        elem.append(para)
        return True

    # Figures: ensure they have mediaobject
    if tag in {'figure', 'informalfigure'}:
        if elem.find('mediaobject') is None and elem.find('graphic') is None:
            mediaobject = etree.SubElement(elem, 'mediaobject')
            textobject = etree.SubElement(mediaobject, 'textobject')
            phrase = etree.SubElement(textobject, 'phrase')
            phrase.text = '[Missing content]'
            return True

    # Examples
    if tag in {'example', 'informalexample'}:
        para = etree.Element('para')
        elem.append(para)
        return True

    # Media objects
    if tag == 'mediaobject':
        textobject = etree.SubElement(elem, 'textobject')
        phrase = etree.SubElement(textobject, 'phrase')
        phrase.text = '[Missing content]'
        return True

    if tag == 'imageobject':
        imagedata = etree.SubElement(elem, 'imagedata')
        imagedata.set('fileref', 'missing.png')
        return True

    # Table structures - these are harder to fix
    if tag in {'table', 'informaltable', 'tgroup', 'tbody', 'row'}:
        # These require more complex fixes - mark as unfixed
        return False

    # TOC entries
    if tag in {'tocpart', 'tocchap', 'toclevel1', 'toclevel2', 'toclevel3', 'toclevel4', 'toclevel5'}:
        tocentry = etree.SubElement(elem, 'tocentry')
        return True

    # Default: can't auto-fix
    return False


def check_element_ordering(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Check if child elements are in correct order per DTD.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    if tag not in ELEMENT_ORDERING:
        return True

    ordering_rules = ELEMENT_ORDERING[tag]
    all_valid = True

    # Check beginpage position - only for elements that have beginpage BEFORE title per DTD
    # Elements like chapter/preface/appendix/part/subpart/toc/lot have beginpage first
    # Sections (sect1-5, section) have beginpage in divcomponent.mix AFTER title
    has_beginpage_first = any(
        child_tag == 'beginpage' and position == 'first'
        for child_tag, required, position in ordering_rules
    )

    if has_beginpage_first:
        beginpage = elem.find('beginpage')
        title = elem.find('title')

        if beginpage is not None and title is not None:
            children = list(elem)
            bp_idx = children.index(beginpage)
            title_idx = children.index(title)

            if bp_idx > title_idx:
                # Fix: move beginpage before title
                elem.remove(beginpage)
                title_idx = list(elem).index(title)
                elem.insert(title_idx, beginpage)

                issue = ValidationIssue(
                    element_tag=tag,
                    element_id=elem_id,
                    issue_type='element_order',
                    description=f"<beginpage> must come before <title> in <{tag}>",
                    severity='error',
                    auto_fixed=True,
                    fix_description="Moved beginpage before title"
                )
                report.add_issue(issue)

    # Check for required title
    for child_tag, required, position in ordering_rules:
        if required and child_tag == 'title':
            if elem.find('title') is None:
                # Add empty title
                title = etree.Element('title')
                # Insert at appropriate position
                insert_pos = 0
                for i, child in enumerate(elem):
                    if child.tag in {'beginpage', f'{tag}info', 'blockinfo'}:
                        insert_pos = i + 1
                    else:
                        break
                elem.insert(insert_pos, title)

                issue = ValidationIssue(
                    element_tag=tag,
                    element_id=elem_id,
                    issue_type='missing_required',
                    description=f"<{tag}> requires <title> element",
                    severity='error',
                    auto_fixed=True,
                    fix_description="Added empty title element"
                )
                report.add_issue(issue)

    return all_valid


def check_table_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate table structure per CALS table model.

    DTD allows tables to use one of these content models:
    - tgroup+ (CALS table with rows/entries)
    - mediaobject+ (table rendered as image)
    - graphic+ (table rendered as graphic)
    """
    if elem.tag not in {'table', 'informaltable'}:
        return True

    elem_id = elem.get('id', 'unknown')

    # Check if table uses mediaobject or graphic content model
    # These are valid alternatives to tgroup per DTD
    has_media = elem.find('mediaobject') is not None or elem.find('graphic') is not None
    if has_media:
        # Valid table using mediaobject/graphic content model - no tgroup needed
        return True

    # Check for tgroup (required when not using mediaobject/graphic)
    tgroup = elem.find('tgroup')
    if tgroup is None:
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"Table requires <tgroup> element (or mediaobject/graphic)",
            severity='error',
            auto_fixed=False
        )
        report.add_issue(issue)
        return False

    # Check tgroup has tbody
    tbody = tgroup.find('tbody')
    if tbody is None:
        issue = ValidationIssue(
            element_tag='tgroup',
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<tgroup> requires <tbody> element",
            severity='error',
            auto_fixed=False
        )
        report.add_issue(issue)
        return False

    # Check tbody has rows
    rows = tbody.findall('row')
    if not rows:
        issue = ValidationIssue(
            element_tag='tbody',
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<tbody> requires at least one <row>",
            severity='error',
            auto_fixed=False
        )
        report.add_issue(issue)
        return False

    # Check each row has entries
    for row in rows:
        entries = row.findall('entry')
        if not entries:
            # Try to fix by adding empty entry
            entry = etree.SubElement(row, 'entry')
            issue = ValidationIssue(
                element_tag='row',
                element_id=elem_id,
                issue_type='empty_element',
                description=f"<row> requires at least one <entry>",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty entry"
            )
            report.add_issue(issue)

    return True


def check_list_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate list structures.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    if tag == 'itemizedlist' or tag == 'orderedlist':
        listitems = elem.findall('listitem')
        if not listitems:
            # Add empty listitem
            listitem = etree.SubElement(elem, 'listitem')
            para = etree.SubElement(listitem, 'para')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='empty_element',
                description=f"<{tag}> requires at least one <listitem>",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty listitem with para"
            )
            report.add_issue(issue)
        else:
            # Check each listitem has content
            for listitem in listitems:
                valid_children = [c for c in listitem if c.tag not in NON_CONTENT_ELEMENTS]
                if not valid_children:
                    para = etree.SubElement(listitem, 'para')
                    issue = ValidationIssue(
                        element_tag='listitem',
                        element_id=elem_id,
                        issue_type='empty_element',
                        description=f"<listitem> requires content",
                        severity='error',
                        auto_fixed=True,
                        fix_description="Added empty para to listitem"
                    )
                    report.add_issue(issue)
        return True

    if tag == 'variablelist':
        entries = elem.findall('varlistentry')
        if not entries:
            varlistentry = etree.SubElement(elem, 'varlistentry')
            term = etree.SubElement(varlistentry, 'term')
            listitem = etree.SubElement(varlistentry, 'listitem')
            para = etree.SubElement(listitem, 'para')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='empty_element',
                description=f"<variablelist> requires at least one <varlistentry>",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty varlistentry"
            )
            report.add_issue(issue)
        else:
            for entry in entries:
                # Check for term
                if entry.find('term') is None:
                    term = etree.Element('term')
                    entry.insert(0, term)
                # Check for listitem
                listitem = entry.find('listitem')
                if listitem is None:
                    listitem = etree.SubElement(entry, 'listitem')
                    para = etree.SubElement(listitem, 'para')
                else:
                    valid_children = [c for c in listitem if c.tag not in NON_CONTENT_ELEMENTS]
                    if not valid_children:
                        para = etree.SubElement(listitem, 'para')
        return True

    if tag == 'simplelist':
        members = elem.findall('member')
        if not members:
            member = etree.SubElement(elem, 'member')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='empty_element',
                description=f"<simplelist> requires at least one <member>",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty member"
            )
            report.add_issue(issue)
        return True

    return True


def check_figure_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate figure structures.
    """
    if elem.tag not in {'figure', 'informalfigure'}:
        return True

    elem_id = elem.get('id', 'unknown')

    # Figure needs title (informalfigure doesn't)
    if elem.tag == 'figure' and elem.find('title') is None:
        title = etree.Element('title')
        elem.insert(0, title)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<figure> requires <title> element",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty title"
        )
        report.add_issue(issue)

    # Check for mediaobject or graphic
    has_content = (elem.find('mediaobject') is not None or
                   elem.find('graphic') is not None or
                   elem.find('inlinemediaobject') is not None)

    if not has_content:
        mediaobject = etree.SubElement(elem, 'mediaobject')
        textobject = etree.SubElement(mediaobject, 'textobject')
        phrase = etree.SubElement(textobject, 'phrase')
        phrase.text = '[Missing figure content]'
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<{elem.tag}> requires mediaobject or graphic content",
            severity='error',
            auto_fixed=True,
            fix_description="Added placeholder mediaobject"
        )
        report.add_issue(issue)

    return True


def check_bibliography_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate bibliography structures.
    """
    if elem.tag == 'bibliography':
        # Must have bibliodiv+ or (biblioentry|bibliomixed)+
        has_content = (elem.find('bibliodiv') is not None or
                       elem.find('biblioentry') is not None or
                       elem.find('bibliomixed') is not None)
        if not has_content:
            bibliomixed = etree.SubElement(elem, 'bibliomixed')
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem.get('id', 'unknown'),
                issue_type='empty_element',
                description=f"<bibliography> requires bibliodiv or biblioentry/bibliomixed",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty bibliomixed"
            )
            report.add_issue(issue)
        return True

    if elem.tag == 'bibliodiv':
        has_content = (elem.find('biblioentry') is not None or
                       elem.find('bibliomixed') is not None)
        if not has_content:
            bibliomixed = etree.SubElement(elem, 'bibliomixed')
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem.get('id', 'unknown'),
                issue_type='empty_element',
                description=f"<bibliodiv> requires biblioentry or bibliomixed",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty bibliomixed"
            )
            report.add_issue(issue)
        return True

    return True


def check_glossary_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate glossary structures.
    Ensures glossentry elements have IDs for proper cross-referencing.
    """
    if elem.tag == 'glossary':
        has_content = (elem.find('glossdiv') is not None or
                       elem.find('glossentry') is not None)
        if not has_content:
            glossentry = etree.SubElement(elem, 'glossentry')
            # Generate ID for glossentry - use chapter ID from parent if available
            chapter_id = _get_chapter_id_from_ancestors(elem) or 'gl0001'
            glossentry_id = f"{chapter_id}s0001gl0001"
            glossentry.set('id', glossentry_id)
            glossterm = etree.SubElement(glossentry, 'glossterm')
            glossdef = etree.SubElement(glossentry, 'glossdef')
            para = etree.SubElement(glossdef, 'para')
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem.get('id', 'unknown'),
                issue_type='empty_element',
                description=f"<glossary> requires glossdiv or glossentry",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty glossentry with ID"
            )
            report.add_issue(issue)
        return True

    if elem.tag == 'glossdiv':
        # Requires title and glossentry+
        if elem.find('title') is None:
            title = etree.Element('title')
            elem.insert(0, title)
        if elem.find('glossentry') is None:
            glossentry = etree.SubElement(elem, 'glossentry')
            # Generate ID for glossentry
            chapter_id = _get_chapter_id_from_ancestors(elem) or 'gl0001'
            # Count existing glossentries to get unique number
            existing_count = len(list(elem.getparent().iter('glossentry'))) if elem.getparent() is not None else 0
            glossentry_id = f"{chapter_id}s0001gl{existing_count + 1:04d}"
            glossentry.set('id', glossentry_id)
            glossterm = etree.SubElement(glossentry, 'glossterm')
            glossdef = etree.SubElement(glossentry, 'glossdef')
            para = etree.SubElement(glossdef, 'para')
        return True

    if elem.tag == 'glossentry':
        # Ensure glossentry has an ID
        if not elem.get('id'):
            chapter_id = _get_chapter_id_from_ancestors(elem) or 'gl0001'
            # Count existing glossentries to generate unique ID
            parent = elem.getparent()
            if parent is not None:
                siblings = [e for e in parent.iter('glossentry')]
                idx = siblings.index(elem) if elem in siblings else len(siblings)
            else:
                idx = 0
            glossentry_id = f"{chapter_id}s0001gl{idx + 1:04d}"
            elem.set('id', glossentry_id)

        if elem.find('glossterm') is None:
            glossterm = etree.Element('glossterm')
            elem.insert(0, glossterm)
        has_def = (elem.find('glossdef') is not None or
                   elem.find('glosssee') is not None)
        if not has_def:
            glossdef = etree.SubElement(elem, 'glossdef')
            para = etree.SubElement(glossdef, 'para')
        return True

    return True


def _get_chapter_id_from_ancestors(elem: etree.Element) -> Optional[str]:
    """Get chapter ID from ancestor elements."""
    parent = elem.getparent()
    while parent is not None:
        parent_tag = parent.tag.split('}')[-1] if '}' in str(parent.tag) else parent.tag
        if parent_tag in ('chapter', 'appendix', 'preface', 'glossary', 'bibliography', 'index'):
            parent_id = parent.get('id')
            if parent_id:
                # Extract base chapter ID (e.g., 'gl0001' from 'gl0001s0001')
                import re
                match = re.match(r'^([a-z]{2}\d{4})', parent_id)
                if match:
                    return match.group(1)
                return parent_id[:6] if len(parent_id) >= 6 else parent_id
        parent = parent.getparent()
    return None


def check_mediaobject_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate mediaobject structures.
    """
    if elem.tag not in {'mediaobject', 'inlinemediaobject'}:
        return True

    elem_id = elem.get('id', 'unknown')

    # Must have at least one of: videoobject, audioobject, imageobject, textobject
    has_content = (elem.find('videoobject') is not None or
                   elem.find('audioobject') is not None or
                   elem.find('imageobject') is not None or
                   elem.find('textobject') is not None)

    if not has_content:
        textobject = etree.SubElement(elem, 'textobject')
        phrase = etree.SubElement(textobject, 'phrase')
        phrase.text = '[Missing media content]'
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<{elem.tag}> requires at least one media object child",
            severity='error',
            auto_fixed=True,
            fix_description="Added placeholder textobject"
        )
        report.add_issue(issue)

    return True


def check_segmented_list_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate segmented list structures.
    Segmented lists require segtitle+ and seglistitem+.
    """
    if elem.tag != 'segmentedlist':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for segtitle (at least one required)
    segtitles = elem.findall('segtitle')
    if not segtitles:
        segtitle = etree.Element('segtitle')
        # Insert at beginning
        elem.insert(0, segtitle)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<segmentedlist> requires at least one <segtitle>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty segtitle"
        )
        report.add_issue(issue)

    # Check for seglistitem (at least one required)
    seglistitems = elem.findall('seglistitem')
    if not seglistitems:
        seglistitem = etree.SubElement(elem, 'seglistitem')
        seg = etree.SubElement(seglistitem, 'seg')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<segmentedlist> requires at least one <seglistitem>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty seglistitem with seg"
        )
        report.add_issue(issue)
    else:
        # Each seglistitem requires at least one seg
        for seglistitem in seglistitems:
            segs = seglistitem.findall('seg')
            if not segs:
                seg = etree.SubElement(seglistitem, 'seg')
                issue = ValidationIssue(
                    element_tag='seglistitem',
                    element_id=elem_id,
                    issue_type='empty_element',
                    description=f"<seglistitem> requires at least one <seg>",
                    severity='error',
                    auto_fixed=True,
                    fix_description="Added empty seg"
                )
                report.add_issue(issue)

    return True


def check_callout_list_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate callout list structures.
    Callout lists require callout+ elements.
    """
    if elem.tag != 'calloutlist':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for callout (at least one required)
    callouts = elem.findall('callout')
    if not callouts:
        callout = etree.SubElement(elem, 'callout')
        callout.set('arearefs', 'area1')  # Required attribute
        para = etree.SubElement(callout, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<calloutlist> requires at least one <callout>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty callout with para"
        )
        report.add_issue(issue)
    else:
        # Each callout requires content
        for callout in callouts:
            valid_children = [c for c in callout if c.tag not in NON_CONTENT_ELEMENTS]
            if not valid_children:
                para = etree.SubElement(callout, 'para')
                issue = ValidationIssue(
                    element_tag='callout',
                    element_id=elem_id,
                    issue_type='empty_element',
                    description=f"<callout> requires content",
                    severity='error',
                    auto_fixed=True,
                    fix_description="Added empty para to callout"
                )
                report.add_issue(issue)

    return True


def check_invalid_nesting(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Check for invalid element nesting that violates DTD.
    """
    # Figure cannot be inside entry (table cell)
    if elem.tag == 'entry':
        for figure in elem.findall('.//figure'):
            # Move figure content to mediaobject (which is allowed in entry)
            mediaobject = figure.find('mediaobject')
            if mediaobject is not None:
                # Replace figure with just the mediaobject
                parent = figure.getparent()
                if parent is not None:
                    idx = list(parent).index(figure)
                    parent.remove(figure)
                    parent.insert(idx, mediaobject)
                    issue = ValidationIssue(
                        element_tag='entry',
                        element_id=elem.get('id', 'unknown'),
                        issue_type='invalid_nesting',
                        description=f"<figure> not allowed inside <entry>, converted to mediaobject",
                        severity='warning',
                        auto_fixed=True,
                        fix_description="Replaced figure with mediaobject"
                    )
                    report.add_issue(issue)

    # Section elements cannot be inside sidebar
    if elem.tag == 'sidebar':
        for section_tag in ['sect1', 'sect2', 'sect3', 'sect4', 'sect5', 'section']:
            for section in elem.findall(f'.//{section_tag}'):
                # Convert section to bridgehead + content
                title = section.find('title')
                if title is not None:
                    # Create bridgehead with title text
                    bridgehead = etree.Element('bridgehead')
                    bridgehead.text = title.text
                    # Move children after title to parent (sidebar)
                    parent = section.getparent()
                    if parent is not None:
                        idx = list(parent).index(section)
                        parent.remove(section)
                        parent.insert(idx, bridgehead)
                        # Move content after bridgehead
                        for i, child in enumerate(list(section)):
                            if child.tag != 'title' and child.tag not in NON_CONTENT_ELEMENTS:
                                parent.insert(idx + 1 + i, child)
                        issue = ValidationIssue(
                            element_tag='sidebar',
                            element_id=elem.get('id', 'unknown'),
                            issue_type='invalid_nesting',
                            description=f"<{section_tag}> not allowed inside <sidebar>, converted to bridgehead",
                            severity='warning',
                            auto_fixed=True,
                            fix_description="Converted section to bridgehead"
                        )
                        report.add_issue(issue)

    # Section elements cannot be inside indexdiv
    # DTD: indexdiv can contain (title?, (indexdivcomponent.mix)*, indexentry+)
    # indexdivcomponent.mix does NOT include sect1/sect2/etc
    if elem.tag == 'indexdiv':
        for section_tag in ['sect1', 'sect2', 'sect3', 'sect4', 'sect5', 'section']:
            for section in list(elem.findall(f'./{section_tag}')):
                # Unwrap section: move children to indexdiv and remove section
                parent = section.getparent()
                if parent is not None:
                    idx = list(parent).index(section)
                    # Move all children of section to parent (except title which becomes bridgehead)
                    title = section.find('title')
                    insert_offset = 0
                    if title is not None:
                        # Convert title to para with bold emphasis (bridgehead is NOT allowed in indexdiv)
                        # indexdivcomponent.mix allows para but not bridgehead
                        para = etree.Element('para')
                        emphasis = etree.SubElement(para, 'emphasis')
                        emphasis.set('role', 'bold')
                        emphasis.text = title.text or ''
                        # Copy any inline children from title
                        for child in title:
                            emphasis.append(child)
                        if title.tail:
                            para.tail = title.tail
                        parent.insert(idx, para)
                        insert_offset = 1
                    # Move other children
                    for i, child in enumerate(list(section)):
                        if child.tag != 'title':
                            parent.insert(idx + insert_offset + i, child)
                    parent.remove(section)
                    issue = ValidationIssue(
                        element_tag='indexdiv',
                        element_id=elem.get('id', 'unknown'),
                        issue_type='invalid_nesting',
                        description=f"<{section_tag}> not allowed inside <indexdiv>, unwrapped content",
                        severity='warning',
                        auto_fixed=True,
                        fix_description="Unwrapped section content into indexdiv"
                    )
                    report.add_issue(issue)

    # bridgehead cannot be inside indexdiv
    # DTD: indexdiv can contain (title?, (indexdivcomponent.mix)*, indexentry+)
    # indexdivcomponent.mix does NOT include bridgehead
    if elem.tag == 'indexdiv':
        for bridgehead in list(elem.findall('.//bridgehead')):
            parent = bridgehead.getparent()
            if parent is not None:
                idx = list(parent).index(bridgehead)
                # Convert bridgehead to para with emphasis
                para = etree.Element('para')
                emphasis = etree.SubElement(para, 'emphasis')
                emphasis.set('role', 'bold')
                emphasis.text = bridgehead.text or ''
                # Copy any children
                for child in bridgehead:
                    emphasis.append(child)
                para.tail = bridgehead.tail
                parent.remove(bridgehead)
                parent.insert(idx, para)
                issue = ValidationIssue(
                    element_tag='indexdiv',
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_nesting',
                    description="<bridgehead> not allowed inside <indexdiv>, converted to para with emphasis",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Converted bridgehead to para with emphasis"
                )
                report.add_issue(issue)

    # bridgehead cannot be inside abstract
    # DTD: abstract can contain (title?, para.class+) - bridgehead is NOT in para.class
    if elem.tag == 'abstract':
        for bridgehead in list(elem.findall('.//bridgehead')):
            parent = bridgehead.getparent()
            if parent is not None:
                idx = list(parent).index(bridgehead)
                # Convert bridgehead to para with emphasis
                para = etree.Element('para')
                emphasis = etree.SubElement(para, 'emphasis')
                emphasis.set('role', 'bold')
                emphasis.text = bridgehead.text or ''
                # Copy any tail content
                for child in bridgehead:
                    emphasis.append(child)
                para.tail = bridgehead.tail
                parent.remove(bridgehead)
                parent.insert(idx, para)
                issue = ValidationIssue(
                    element_tag='abstract',
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_nesting',
                    description="<bridgehead> not allowed inside <abstract>, converted to para with emphasis",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Converted bridgehead to para with emphasis"
                )
                report.add_issue(issue)

    # para cannot be inside bridgehead
    # DTD: bridgehead can only contain title.char.mix (inline content)
    if elem.tag == 'bridgehead':
        for para in list(elem.findall('.//para')):
            parent = para.getparent()
            if parent is not None:
                # Flatten para: extract text and inline content
                # Add para's text content directly to parent
                if para.text:
                    if parent == elem:
                        # Direct child of bridgehead
                        if elem.text:
                            elem.text += ' ' + para.text
                        else:
                            elem.text = para.text
                    else:
                        # Nested deeper - prepend to previous sibling's tail or parent text
                        prev = para.getprevious()
                        if prev is not None:
                            if prev.tail:
                                prev.tail += ' ' + para.text
                            else:
                                prev.tail = para.text
                        elif parent.text:
                            parent.text += ' ' + para.text
                        else:
                            parent.text = para.text
                # Move inline children up
                idx = list(parent).index(para)
                for i, child in enumerate(list(para)):
                    parent.insert(idx + i, child)
                # Preserve tail
                if para.tail:
                    last_child = para[-1] if len(para) > 0 else None
                    if last_child is not None:
                        if last_child.tail:
                            last_child.tail += para.tail
                        else:
                            last_child.tail = para.tail
                    else:
                        # No children moved, append tail to parent text
                        if parent.text:
                            parent.text += para.tail
                        else:
                            parent.text = para.tail
                parent.remove(para)
                issue = ValidationIssue(
                    element_tag='bridgehead',
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_nesting',
                    description="<para> not allowed inside <bridgehead>, flattened content",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Flattened para content into bridgehead"
                )
                report.add_issue(issue)

    # para cannot be inside subtitle
    # DTD: subtitle can only contain title.char.mix (inline content)
    if elem.tag == 'subtitle':
        for para in list(elem.findall('.//para')):
            parent = para.getparent()
            if parent is not None:
                # Flatten para: extract text and inline content
                if para.text:
                    if parent == elem:
                        if elem.text:
                            elem.text += ' ' + para.text
                        else:
                            elem.text = para.text
                    else:
                        prev = para.getprevious()
                        if prev is not None:
                            if prev.tail:
                                prev.tail += ' ' + para.text
                            else:
                                prev.tail = para.text
                        elif parent.text:
                            parent.text += ' ' + para.text
                        else:
                            parent.text = para.text
                # Move inline children up
                idx = list(parent).index(para)
                for i, child in enumerate(list(para)):
                    parent.insert(idx + i, child)
                # Preserve tail
                if para.tail:
                    last_child = para[-1] if len(para) > 0 else None
                    if last_child is not None:
                        if last_child.tail:
                            last_child.tail += para.tail
                        else:
                            last_child.tail = para.tail
                    else:
                        if parent.text:
                            parent.text += para.tail
                        else:
                            parent.text = para.tail
                parent.remove(para)
                issue = ValidationIssue(
                    element_tag='subtitle',
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_nesting',
                    description="<para> not allowed inside <subtitle>, flattened content",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Flattened para content into subtitle"
                )
                report.add_issue(issue)

    # Nested table elements - table cannot contain another table directly
    # DTD: table contains (blockinfo?, title, titleabbrev?, indexterm*, textobject*, tgroup+)
    # A child <table> is NOT allowed - this happens with nested tables
    if elem.tag in ('table', 'informaltable'):
        for nested_table in list(elem.findall('./table')) + list(elem.findall('./informaltable')):
            # Unwrap nested table: extract its tgroup(s) and add to parent
            parent = nested_table.getparent()
            if parent is not None and parent.tag in ('table', 'informaltable'):
                idx = list(parent).index(nested_table)
                # Move tgroups from nested table to parent
                tgroups = nested_table.findall('tgroup')
                for i, tgroup in enumerate(tgroups):
                    parent.insert(idx + i, tgroup)
                # Remove the nested table
                parent.remove(nested_table)
                issue = ValidationIssue(
                    element_tag=elem.tag,
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_nesting',
                    description=f"<{nested_table.tag}> not allowed inside <{elem.tag}>, unwrapped tgroups",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Unwrapped nested table tgroups"
                )
                report.add_issue(issue)

        # Also fix para elements directly inside table (not allowed)
        # DTD: table contains (blockinfo?, title, titleabbrev?, indexterm*, textobject*, tgroup+)
        for para in list(elem.findall('./para')):
            parent = elem.getparent()
            if parent is not None:
                # Move para to after the table
                table_idx = list(parent).index(elem)
                elem.remove(para)
                parent.insert(table_idx + 1, para)
                issue = ValidationIssue(
                    element_tag=elem.tag,
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_nesting',
                    description=f"<para> not allowed inside <{elem.tag}>, moved after table",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Moved para element to after table"
                )
                report.add_issue(issue)

    # indexdiv must end with indexentry+ or segmentedlist (not optional)
    # DTD: indexdiv can contain (title?, (indexdivcomponent.mix)*, (indexentry+ | segmentedlist))
    if elem.tag == 'indexdiv':
        has_indexentry = any(c.tag == 'indexentry' for c in elem)
        has_segmentedlist = any(c.tag == 'segmentedlist' for c in elem)

        if not has_indexentry and not has_segmentedlist:
            # Try to convert itemizedlist to indexentry if present
            itemizedlist = elem.find('itemizedlist')
            if itemizedlist is not None:
                for listitem in list(itemizedlist.findall('listitem')):
                    indexentry = etree.Element('indexentry')
                    primaryie = etree.SubElement(indexentry, 'primaryie')
                    para = listitem.find('para')
                    if para is not None:
                        primaryie.text = ''.join(para.itertext()).strip() or 'Index entry'
                    else:
                        primaryie.text = ''.join(listitem.itertext()).strip() or 'Index entry'
                    elem.append(indexentry)
                elem.remove(itemizedlist)
                issue = ValidationIssue(
                    element_tag='indexdiv',
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_content_model',
                    description="<indexdiv> requires indexentry+ or segmentedlist, converted itemizedlist",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Converted itemizedlist to indexentry elements"
                )
                report.add_issue(issue)
            else:
                # Create minimal placeholder indexentry
                indexentry = etree.Element('indexentry')
                primaryie = etree.SubElement(indexentry, 'primaryie')
                existing_text = ''.join(elem.itertext()).strip()
                primaryie.text = existing_text[:50].strip() if existing_text else 'Index'
                elem.append(indexentry)
                issue = ValidationIssue(
                    element_tag='indexdiv',
                    element_id=elem.get('id', 'unknown'),
                    issue_type='invalid_content_model',
                    description="<indexdiv> requires indexentry+ or segmentedlist, added placeholder",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Added placeholder indexentry element"
                )
                report.add_issue(issue)

    return True


def check_textobject_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate textobject structures.
    textobject requires either objectinfo OR (phrase|textdata|other content)+
    """
    if elem.tag != 'textobject':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for content
    has_content = (elem.find('objectinfo') is not None or
                   elem.find('phrase') is not None or
                   elem.find('textdata') is not None or
                   len(list(elem)) > 0 or
                   (elem.text and elem.text.strip()))

    if not has_content:
        phrase = etree.SubElement(elem, 'phrase')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<textobject> requires content (phrase, textdata, or other)",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty phrase"
        )
        report.add_issue(issue)

    return True


def check_required_attributes(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate that required attributes are present on elements.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    # tgroup requires cols attribute
    if tag == 'tgroup':
        if not elem.get('cols'):
            # Try to auto-fix by counting colspec or entry elements
            colspecs = elem.findall('colspec')
            if colspecs:
                cols = len(colspecs)
            else:
                # Count entries in first row
                tbody = elem.find('tbody')
                if tbody is not None:
                    first_row = tbody.find('row')
                    if first_row is not None:
                        cols = len(first_row.findall('entry'))
                    else:
                        cols = 1
                else:
                    cols = 1
            elem.set('cols', str(cols))
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<tgroup> requires 'cols' attribute",
                severity='error',
                auto_fixed=True,
                fix_description=f"Set cols='{cols}' based on structure"
            )
            report.add_issue(issue)

    # entrytbl requires cols attribute
    if tag == 'entrytbl':
        if not elem.get('cols'):
            elem.set('cols', '1')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<entrytbl> requires 'cols' attribute",
                severity='error',
                auto_fixed=True,
                fix_description="Set cols='1' as default"
            )
            report.add_issue(issue)

    # spanspec requires namest, nameend, spanname attributes
    if tag == 'spanspec':
        missing = []
        if not elem.get('namest'):
            missing.append('namest')
        if not elem.get('nameend'):
            missing.append('nameend')
        if not elem.get('spanname'):
            missing.append('spanname')
        if missing:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<spanspec> requires attributes: {', '.join(missing)}",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)

    # callout requires arearefs attribute
    if tag == 'callout':
        if not elem.get('arearefs'):
            elem.set('arearefs', 'area1')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<callout> requires 'arearefs' attribute",
                severity='error',
                auto_fixed=True,
                fix_description="Set arearefs='area1' as placeholder"
            )
            report.add_issue(issue)

    # xref requires linkend attribute
    if tag == 'xref':
        if not elem.get('linkend'):
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<xref> requires 'linkend' attribute",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)

    # link requires linkend (for internal links) or check for url/xlink:href
    if tag == 'link':
        if not elem.get('linkend') and not elem.get('url') and not elem.get('{http://www.w3.org/1999/xlink}href'):
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<link> requires 'linkend' or 'url' attribute",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # ulink requires url attribute
    if tag == 'ulink':
        if not elem.get('url'):
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<ulink> requires 'url' attribute",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)

    # anchor requires id attribute
    if tag == 'anchor':
        if not elem.get('id'):
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<anchor> requires 'id' attribute",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)

    # area requires coords attribute
    if tag == 'area':
        if not elem.get('coords'):
            elem.set('coords', '0,0')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<area> requires 'coords' attribute",
                severity='error',
                auto_fixed=True,
                fix_description="Set coords='0,0' as placeholder"
            )
            report.add_issue(issue)

    # co (callout mark) requires id attribute
    if tag == 'co':
        if not elem.get('id'):
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<co> requires 'id' attribute",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)

    return True


def check_formalpara_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate formalpara structure: requires title followed by para.
    Content model: (title, indexterm*, para)
    """
    if elem.tag != 'formalpara':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for required title
    title = elem.find('title')
    if title is None:
        title = etree.Element('title')
        elem.insert(0, title)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<formalpara> requires <title> element",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty title"
        )
        report.add_issue(issue)

    # Check for required para
    para = elem.find('para')
    if para is None:
        para = etree.SubElement(elem, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<formalpara> requires <para> element",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty para"
        )
        report.add_issue(issue)

    # Check ordering: title must come before para
    children = list(elem)
    title_idx = children.index(title) if title in children else -1
    para_idx = children.index(para) if para in children else -1

    if title_idx > para_idx and para_idx >= 0:
        # Move title before para
        elem.remove(title)
        elem.insert(para_idx, title)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='element_order',
            description=f"<title> must come before <para> in <formalpara>",
            severity='error',
            auto_fixed=True,
            fix_description="Moved title before para"
        )
        report.add_issue(issue)

    return True


def check_procedure_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate procedure structure: requires step+.
    Content model: (blockinfo?, title?, component.mix*, step+)
    """
    if elem.tag != 'procedure':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for at least one step
    steps = elem.findall('step')
    if not steps:
        step = etree.SubElement(elem, 'step')
        para = etree.SubElement(step, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<procedure> requires at least one <step>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty step with para"
        )
        report.add_issue(issue)

    return True


def check_substeps_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate substeps structure: requires step+.
    """
    if elem.tag not in {'substeps', 'stepalternatives'}:
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for at least one step
    steps = elem.findall('step')
    if not steps:
        step = etree.SubElement(elem, 'step')
        para = etree.SubElement(step, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<{elem.tag}> requires at least one <step>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty step with para"
        )
        report.add_issue(issue)

    return True


def check_refentry_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate refentry structure: requires refnamediv+.
    Content model: (beginpage?, indexterm*, refentryinfo?, refmeta?,
                    (remark|link.char.class)*, refnamediv+,
                    refsynopsisdiv?, (refsect1+|refsection+))
    """
    if elem.tag != 'refentry':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for at least one refnamediv
    refnamedivs = elem.findall('refnamediv')
    if not refnamedivs:
        refnamediv = etree.SubElement(elem, 'refnamediv')
        refname = etree.SubElement(refnamediv, 'refname')
        refpurpose = etree.SubElement(refnamediv, 'refpurpose')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<refentry> requires at least one <refnamediv>",
            severity='error',
            auto_fixed=True,
            fix_description="Added refnamediv with refname and refpurpose"
        )
        report.add_issue(issue)

    return True


def check_refnamediv_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate refnamediv structure: requires refname+ and refpurpose.
    Content model: (refdescriptor?, refname+, refpurpose, refclass*)
    """
    if elem.tag != 'refnamediv':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for at least one refname
    refnames = elem.findall('refname')
    if not refnames:
        refname = etree.SubElement(elem, 'refname')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<refnamediv> requires at least one <refname>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty refname"
        )
        report.add_issue(issue)

    # Check for refpurpose
    refpurpose = elem.find('refpurpose')
    if refpurpose is None:
        refpurpose = etree.SubElement(elem, 'refpurpose')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<refnamediv> requires <refpurpose>",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty refpurpose"
        )
        report.add_issue(issue)

    return True


def check_info_element_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate info elements: must not be empty if present.
    Info elements like chapterinfo, sect1info, etc. require info.class+ children.
    """
    info_tags = {
        'chapterinfo', 'appendixinfo', 'prefaceinfo', 'bookinfo',
        'sect1info', 'sect2info', 'sect3info', 'sect4info', 'sect5info',
        'sectioninfo', 'articleinfo', 'partinfo', 'setinfo',
        'refentryinfo', 'refsect1info', 'refsect2info', 'refsect3info',
        'blockinfo', 'objectinfo', 'sidebarinfo', 'bibliographyinfo',
        'glossaryinfo', 'indexinfo', 'referenceinfo'
    }

    if elem.tag not in info_tags:
        return True

    elem_id = elem.get('id', 'unknown')

    # Check if info element has any children
    if len(list(elem)) == 0 and not (elem.text and elem.text.strip()):
        # Info element is empty - remove it (it's optional, better to remove than have invalid)
        parent = elem.getparent()
        if parent is not None:
            parent.remove(elem)
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem_id,
                issue_type='empty_element',
                description=f"<{elem.tag}> cannot be empty, removed",
                severity='warning',
                auto_fixed=True,
                fix_description="Removed empty info element"
            )
            report.add_issue(issue)

    return True


def check_imageobject_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate imageobject/videoobject/audioobject: require data element.
    """
    data_requirements = {
        'imageobject': 'imagedata',
        'videoobject': 'videodata',
        'audioobject': 'audiodata',
    }

    if elem.tag not in data_requirements:
        return True

    required_data = data_requirements[elem.tag]
    elem_id = elem.get('id', 'unknown')

    data_elem = elem.find(required_data)
    if data_elem is None:
        # Add the required data element
        data_elem = etree.SubElement(elem, required_data)
        data_elem.set('fileref', 'missing.dat')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<{elem.tag}> requires <{required_data}> element",
            severity='error',
            auto_fixed=True,
            fix_description=f"Added {required_data} with placeholder fileref"
        )
        report.add_issue(issue)

    return True


def _update_linkend_references(root: etree.Element, old_id: str, new_id: str) -> int:
    """
    Update all linkend and url attributes referencing old_id to new_id.

    Also updates child element IDs that use the old section ID as a prefix.
    """
    updated = 0
    for elem in root.iter():
        # Update linkend attributes
        linkend = elem.get('linkend')
        if linkend == old_id:
            elem.set('linkend', new_id)
            updated += 1
        elif linkend and linkend.startswith(old_id):
            # Update linkends that reference elements within this section
            # e.g., ch0005s0001s10006fg01 -> ch0005s0006fg01
            new_linkend = new_id + linkend[len(old_id):]
            elem.set('linkend', new_linkend)
            updated += 1

        # Update url attributes (for ulink elements)
        url = elem.get('url')
        if url == old_id:
            elem.set('url', new_id)
            updated += 1
        elif url and old_id in url:
            new_url = url.replace(old_id, new_id)
            elem.set('url', new_url)
            updated += 1

    return updated


def _update_child_element_ids(section_elem: etree.Element, old_sect_id: str, new_sect_id: str) -> int:
    """
    Update IDs of child elements to use the new section ID prefix.

    When a section's ID changes (e.g., from sect2 format to sect1 format),
    child element IDs that used the old section as a prefix must be updated.
    """
    updated = 0
    for elem in section_elem.iter():
        if elem == section_elem:
            continue  # Skip the section element itself
        elem_id = elem.get('id')
        if elem_id and elem_id.startswith(old_sect_id):
            new_elem_id = new_sect_id + elem_id[len(old_sect_id):]
            elem.set('id', new_elem_id)
            updated += 1
            logger.debug(f"Updated child element ID: {elem_id} -> {new_elem_id}")
    return updated


def check_section_nesting(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate section nesting: sect2 can only appear in sect1, etc.

    When converting a section to a different level (e.g., sect2 to sect1),
    the ID is regenerated to match the new section format to ensure
    downstream XSL processing works correctly.
    """
    nesting_rules = {
        'sect2': {'sect1'},
        'sect3': {'sect2'},
        'sect4': {'sect3'},
        'sect5': {'sect4'},
    }

    if elem.tag not in nesting_rules:
        return True

    allowed_parents = nesting_rules[elem.tag]
    parent = elem.getparent()

    if parent is not None and parent.tag not in allowed_parents:
        # This is a nesting violation - convert to appropriate level or bridgehead
        old_id = elem.get('id', '')
        original_tag = elem.tag

        # Find correct section level based on parent
        if parent.tag in {'chapter', 'appendix', 'preface', 'dedication', 'partintro'}:
            # Convert to sect1 and regenerate ID
            elem.tag = 'sect1'

            # Get parent ID for generating new sect1 ID
            parent_id = parent.get('id', '')
            if not parent_id:
                # Try to extract from filename pattern in ID
                id_match = re.match(r'^([a-z]{2}\d{4})', old_id)
                parent_id = id_match.group(1) if id_match else 'ch0001'

            # Find root element to collect existing IDs
            root = elem
            while root.getparent() is not None:
                root = root.getparent()

            existing_ids = {e.get('id') for e in root.iter() if e.get('id')}

            # Generate new compliant sect1 ID
            new_id = next_available_sect1_id(parent_id, existing_ids)
            elem.set('id', new_id)

            # Update child element IDs to use new section prefix
            child_updates = _update_child_element_ids(elem, old_id, new_id) if old_id else 0

            # Update linkend references throughout document
            linkend_updates = _update_linkend_references(root, old_id, new_id) if old_id else 0

            fix_desc = f"Converted to sect1, ID changed from '{old_id}' to '{new_id}'"
            if child_updates > 0:
                fix_desc += f", updated {child_updates} child IDs"
            if linkend_updates > 0:
                fix_desc += f", updated {linkend_updates} linkend refs"

            issue = ValidationIssue(
                element_tag=original_tag,
                element_id=new_id,
                issue_type='invalid_nesting',
                description=f"<{original_tag}> found directly in <{parent.tag}>, must be in <sect1>",
                severity='warning',
                auto_fixed=True,
                fix_description=fix_desc
            )
            report.add_issue(issue)
            logger.info(f"Converted <{original_tag} id='{old_id}'> to <sect1 id='{new_id}'> in <{parent.tag}>")

        elif parent.tag == 'sect1' and elem.tag == 'sect3':
            # sect3 in sect1 - convert to sect2
            # Keep the ID since it's still within the same section hierarchy
            elem.tag = 'sect2'
            issue = ValidationIssue(
                element_tag='sect3',
                element_id=old_id,
                issue_type='invalid_nesting',
                description=f"<sect3> found directly in <sect1>, must be in <sect2>",
                severity='warning',
                auto_fixed=True,
                fix_description="Converted to sect2"
            )
            report.add_issue(issue)

    return True


def check_refsect_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate refsect1/refsect2/refsect3: require title.
    Unlike regular sect*, title is REQUIRED in refsect*.
    """
    if elem.tag not in {'refsect1', 'refsect2', 'refsect3', 'refsection'}:
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for required title
    title = elem.find('title')
    if title is None:
        # Find correct insertion point (after *info if present)
        insert_pos = 0
        info_tag = f"{elem.tag}info"
        for i, child in enumerate(elem):
            if child.tag == info_tag:
                insert_pos = i + 1
                break

        title = etree.Element('title')
        elem.insert(insert_pos, title)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<{elem.tag}> requires <title> element",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty title"
        )
        report.add_issue(issue)

    return True


def check_abstract_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate abstract: requires para.class+ (at least one paragraph).
    Content model: (title?, para.class+)
    """
    if elem.tag != 'abstract':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for at least one para/simpara/formalpara
    para_elements = ['para', 'simpara', 'formalpara']
    has_para = any(elem.find(p) is not None for p in para_elements)

    if not has_para:
        para = etree.SubElement(elem, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<abstract> requires at least one paragraph element",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty para"
        )
        report.add_issue(issue)

    return True


def check_answer_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate answer: requires qandaset.mix+ content.
    Content model: (label?, qandaset.mix+, qandaentry*)
    """
    if elem.tag != 'answer':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for content (excluding label and qandaentry)
    valid_children = [c for c in elem if c.tag not in {'label', 'qandaentry'}]
    if not valid_children:
        para = etree.SubElement(elem, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<answer> requires content",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty para"
        )
        report.add_issue(issue)

    return True


def check_qandaentry_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate qandaentry: requires question element.
    Content model: (blockinfo?, revhistory?, question, answer*)
    """
    if elem.tag != 'qandaentry':
        return True

    elem_id = elem.get('id', 'unknown')

    # Check for required question
    question = elem.find('question')
    if question is None:
        question = etree.SubElement(elem, 'question')
        para = etree.SubElement(question, 'para')
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<qandaentry> requires <question> element",
            severity='error',
            auto_fixed=True,
            fix_description="Added question with empty para"
        )
        report.add_issue(issue)

    return True


def check_tgroup_ordering(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate tgroup element ordering: colspec*, spanspec*, thead?, tfoot?, tbody.
    Elements must appear in this order.
    """
    if elem.tag != 'tgroup':
        return True

    elem_id = elem.get('id', 'unknown')
    children = list(elem)

    # Define expected order
    order_map = {
        'colspec': 0,
        'spanspec': 1,
        'thead': 2,
        'tfoot': 3,
        'tbody': 4,
    }

    last_order = -1
    needs_reorder = False

    for child in children:
        if child.tag in order_map:
            current_order = order_map[child.tag]
            if current_order < last_order:
                needs_reorder = True
                break
            last_order = current_order

    if needs_reorder:
        # Reorder children according to DTD
        new_children = []
        remaining = list(elem)
        for elem_tag in ['colspec', 'spanspec', 'thead', 'tfoot', 'tbody']:
            for child in remaining[:]:
                if child.tag == elem_tag:
                    new_children.append(child)
                    remaining.remove(child)
        # Add any remaining children at the end
        new_children.extend(remaining)

        # Clear and re-add in correct order
        for child in list(elem):
            elem.remove(child)
        for child in new_children:
            elem.append(child)

        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='element_order',
            description=f"<tgroup> children must be in order: colspec*, spanspec*, thead?, tfoot?, tbody",
            severity='error',
            auto_fixed=True,
            fix_description="Reordered tgroup children"
        )
        report.add_issue(issue)

    return True


# =============================================================================
# HIGH PRIORITY: REFERENTIAL INTEGRITY CHECKS
# =============================================================================

def check_referential_integrity(root_elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate referential integrity: ensure all ID references point to existing IDs.
    This is a document-wide check that runs after all elements are processed.
    """
    # First, collect all IDs in the document
    all_ids = set()
    for elem in root_elem.iter():
        if isinstance(elem.tag, str):
            elem_id = elem.get('id')
            if elem_id:
                all_ids.add(elem_id)

    # Now check all references
    issues_found = 0

    for elem in root_elem.iter():
        if not isinstance(elem.tag, str):
            continue

        tag = elem.tag

        # xref/@linkend must point to existing ID
        if tag == 'xref':
            linkend = elem.get('linkend')
            if linkend and linkend not in all_ids:
                issue = ValidationIssue(
                    element_tag=tag,
                    element_id=elem.get('id', 'unknown'),
                    issue_type='broken_reference',
                    description=f"<xref linkend='{linkend}'> references non-existent ID",
                    severity='error',
                    auto_fixed=False
                )
                report.add_issue(issue)
                issues_found += 1

        # link/@linkend must point to existing ID (if using linkend)
        if tag == 'link':
            linkend = elem.get('linkend')
            if linkend and linkend not in all_ids:
                issue = ValidationIssue(
                    element_tag=tag,
                    element_id=elem.get('id', 'unknown'),
                    issue_type='broken_reference',
                    description=f"<link linkend='{linkend}'> references non-existent ID",
                    severity='error',
                    auto_fixed=False
                )
                report.add_issue(issue)
                issues_found += 1

        # footnoteref/@linkend must point to existing footnote ID
        if tag == 'footnoteref':
            linkend = elem.get('linkend')
            if linkend and linkend not in all_ids:
                issue = ValidationIssue(
                    element_tag=tag,
                    element_id=elem.get('id', 'unknown'),
                    issue_type='broken_reference',
                    description=f"<footnoteref linkend='{linkend}'> references non-existent footnote",
                    severity='error',
                    auto_fixed=False
                )
                report.add_issue(issue)
                issues_found += 1

        # callout/@arearefs must point to existing area IDs (space-separated IDREFS)
        if tag == 'callout':
            arearefs = elem.get('arearefs')
            if arearefs:
                for arearef in arearefs.split():
                    if arearef not in all_ids:
                        issue = ValidationIssue(
                            element_tag=tag,
                            element_id=elem.get('id', 'unknown'),
                            issue_type='broken_reference',
                            description=f"<callout arearefs> references non-existent area '{arearef}'",
                            severity='warning',
                            auto_fixed=False
                        )
                        report.add_issue(issue)
                        issues_found += 1

        # TOC elements: tocentry/@linkend, tocfront/@linkend, tocback/@linkend
        if tag in {'tocentry', 'tocfront', 'tocback'}:
            linkend = elem.get('linkend')
            if linkend and linkend not in all_ids:
                issue = ValidationIssue(
                    element_tag=tag,
                    element_id=elem.get('id', 'unknown'),
                    issue_type='broken_reference',
                    description=f"<{tag} linkend='{linkend}'> references non-existent ID",
                    severity='error',
                    auto_fixed=False
                )
                report.add_issue(issue)
                issues_found += 1

        # indexterm with class="endofrange" must have startref pointing to existing ID
        if tag == 'indexterm':
            idx_class = elem.get('class')
            startref = elem.get('startref')
            if idx_class == 'endofrange':
                if not startref:
                    issue = ValidationIssue(
                        element_tag=tag,
                        element_id=elem.get('id', 'unknown'),
                        issue_type='missing_required_attr',
                        description=f"<indexterm class='endofrange'> requires 'startref' attribute",
                        severity='error',
                        auto_fixed=False
                    )
                    report.add_issue(issue)
                    issues_found += 1
                elif startref not in all_ids:
                    issue = ValidationIssue(
                        element_tag=tag,
                        element_id=elem.get('id', 'unknown'),
                        issue_type='broken_reference',
                        description=f"<indexterm startref='{startref}'> references non-existent ID",
                        severity='error',
                        auto_fixed=False
                    )
                    report.add_issue(issue)
                    issues_found += 1

    if issues_found > 0:
        logger.warning(f"Found {issues_found} broken references in document")

    return issues_found == 0


# =============================================================================
# HIGH PRIORITY: CONTENT MODEL RESTRICTIONS BY PARENT CONTEXT
# =============================================================================

def check_content_restrictions(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate that elements only contain allowed children based on their context.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    # entry (table cell) content restrictions
    # entry can contain: para, list, mediaobject, admonitions, programlisting, etc.
    # entry CANNOT contain: figure, table, example, sidebar (formal objects)
    if tag == 'entry':
        forbidden_in_entry = {'figure', 'table', 'informaltable', 'example',
                              'informalexample', 'sidebar', 'equation', 'informalequation'}
        for forbidden in forbidden_in_entry:
            for child in elem.findall(f'.//{forbidden}'):
                # Try to fix: convert figure to mediaobject, remove others
                if forbidden == 'figure':
                    mediaobject = child.find('mediaobject')
                    if mediaobject is not None:
                        parent = child.getparent()
                        if parent is not None:
                            idx = list(parent).index(child)
                            parent.remove(child)
                            parent.insert(idx, mediaobject)
                            issue = ValidationIssue(
                                element_tag='entry',
                                element_id=elem_id,
                                issue_type='invalid_content',
                                description=f"<figure> not allowed in <entry>, extracted mediaobject",
                                severity='warning',
                                auto_fixed=True,
                                fix_description="Replaced figure with its mediaobject"
                            )
                            report.add_issue(issue)
                else:
                    issue = ValidationIssue(
                        element_tag='entry',
                        element_id=elem_id,
                        issue_type='invalid_content',
                        description=f"<{forbidden}> not allowed inside <entry>",
                        severity='error',
                        auto_fixed=False
                    )
                    report.add_issue(issue)

    # sidebar content restrictions
    # sidebar CANNOT contain: sect1-sect5, section (sections not allowed)
    # Already handled in check_invalid_nesting, but double-check formal objects
    if tag == 'sidebar':
        forbidden_in_sidebar = {'figure', 'table', 'example', 'equation'}
        for forbidden in forbidden_in_sidebar:
            for child in elem.findall(f'.//{forbidden}'):
                issue = ValidationIssue(
                    element_tag='sidebar',
                    element_id=elem_id,
                    issue_type='invalid_content',
                    description=f"<{forbidden}> (formal object) not allowed inside <sidebar>",
                    severity='error',
                    auto_fixed=False
                )
                report.add_issue(issue)

    # glossdef content restrictions
    # glossdef CANNOT contain: anchor, bridgehead, highlights
    if tag == 'glossdef':
        forbidden_in_glossdef = {'anchor', 'bridgehead', 'highlights'}
        for forbidden in forbidden_in_glossdef:
            for child in elem.findall(f'.//{forbidden}'):
                if forbidden == 'anchor':
                    # Remove anchor (not critical, just not allowed here)
                    parent = child.getparent()
                    if parent is not None:
                        parent.remove(child)
                        issue = ValidationIssue(
                            element_tag='glossdef',
                            element_id=elem_id,
                            issue_type='invalid_content',
                            description=f"<anchor> not allowed inside <glossdef>, removed",
                            severity='warning',
                            auto_fixed=True,
                            fix_description="Removed anchor from glossdef"
                        )
                        report.add_issue(issue)
                else:
                    issue = ValidationIssue(
                        element_tag='glossdef',
                        element_id=elem_id,
                        issue_type='invalid_content',
                        description=f"<{forbidden}> not allowed inside <glossdef>",
                        severity='error',
                        auto_fixed=False
                    )
                    report.add_issue(issue)

    # legalnotice content restrictions
    # legalnotice has restricted content - only specific block elements
    if tag == 'legalnotice':
        forbidden_in_legalnotice = {'graphic', 'mediaobject', 'informalfigure', 'figure'}
        for forbidden in forbidden_in_legalnotice:
            for child in elem.findall(f'.//{forbidden}'):
                issue = ValidationIssue(
                    element_tag='legalnotice',
                    element_id=elem_id,
                    issue_type='invalid_content',
                    description=f"<{forbidden}> not allowed inside <legalnotice>",
                    severity='warning',
                    auto_fixed=False
                )
                report.add_issue(issue)

    return True


# =============================================================================
# HIGH PRIORITY: MIXED CONTENT PCDATA DETECTION
# =============================================================================

def has_text_content(elem: etree.Element) -> bool:
    """
    Check if an element has meaningful text content (PCDATA).
    Returns True if the element or any descendant has non-whitespace text.
    """
    # Check direct text
    if elem.text and elem.text.strip():
        return True

    # Check tail text of children
    for child in elem:
        if child.tail and child.tail.strip():
            return True
        # Recursively check children
        if has_text_content(child):
            return True

    return False


# Elements that can have mixed content (PCDATA with elements)
MIXED_CONTENT_ELEMENTS = {
    'para', 'simpara', 'programlisting', 'literallayout', 'screen',
    'synopsis', 'address', 'title', 'subtitle', 'titleabbrev',
    'term', 'member', 'seg', 'primary', 'secondary', 'tertiary',
    'see', 'seealso', 'refname', 'refpurpose', 'refdescriptor',
    'phrase', 'emphasis', 'literal', 'code', 'command', 'option',
    'computeroutput', 'userinput', 'filename', 'classname', 'methodname',
    'varname', 'funcdef', 'paramdef', 'parameter', 'returnvalue',
    'type', 'structname', 'structfield', 'symbol', 'token',
    'markup', 'prompt', 'envar', 'errorcode', 'errorname', 'errortype',
    'caption', 'attribution', 'bibliomixed', 'bibliomset',
    'lineannotation', 'remark', 'funcsynopsisinfo', 'classsynopsisinfo',
}


def check_mixed_content_empty(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    For mixed content elements, check if they have PCDATA before flagging as empty.
    This prevents false positives for elements like <para>Just text</para>.
    """
    if elem.tag not in MIXED_CONTENT_ELEMENTS:
        return True

    # If element has text content, it's not empty even without child elements
    if has_text_content(elem):
        return True

    # Element has no text and no children - may need content
    # But some mixed content elements can be empty (like empty para for spacing)
    # Only flag if it's in a context where content is required

    return True


# =============================================================================
# MEDIUM PRIORITY: CHOICE GROUP CONSTRAINTS
# =============================================================================

def check_choice_group_constraints(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate choice group constraints where exactly one of several options must be present.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    # glossentry: must have EITHER glossdef OR glosssee (at least one, can have both)
    if tag == 'glossentry':
        has_glossdef = elem.find('glossdef') is not None
        has_glosssee = elem.find('glosssee') is not None

        if not has_glossdef and not has_glosssee:
            # Add glossdef as default
            glossdef = etree.SubElement(elem, 'glossdef')
            para = etree.SubElement(glossdef, 'para')
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required',
                description=f"<glossentry> requires either <glossdef> or <glosssee>",
                severity='error',
                auto_fixed=True,
                fix_description="Added empty glossdef with para"
            )
            report.add_issue(issue)

    # refentry: should have refsect1+ OR refsection+ (not mixed)
    if tag == 'refentry':
        has_refsect1 = elem.find('refsect1') is not None
        has_refsection = elem.find('refsection') is not None

        if has_refsect1 and has_refsection:
            # Mixed usage - convert refsection to refsect1
            for refsection in elem.findall('refsection'):
                refsection.tag = 'refsect1'
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_content',
                description=f"<refentry> should use either refsect1 OR refsection, not both",
                severity='warning',
                auto_fixed=True,
                fix_description="Converted refsection to refsect1"
            )
            report.add_issue(issue)

    # mediaobject: must have at least one of videoobject|audioobject|imageobject|textobject
    # (Already handled in check_mediaobject_structure)

    # equation/informalequation: must have content
    if tag in {'equation', 'informalequation'}:
        # Check for valid equation content - DTD: (alt?, (graphic+|mediaobject+))
        valid_content = {'mediaobject', 'graphic', 'alt'}
        has_content = any(elem.find(c) is not None for c in valid_content)
        if not has_content and not has_text_content(elem):
            # Add placeholder mediaobject with textobject
            mediaobj = etree.SubElement(elem, 'mediaobject')
            textobj = etree.SubElement(mediaobj, 'textobject')
            phrase = etree.SubElement(textobj, 'phrase')
            phrase.text = '[equation]'
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='empty_element',
                description=f"<{tag}> requires equation content",
                severity='error',
                auto_fixed=True,
                fix_description="Added placeholder mediaobject"
            )
            report.add_issue(issue)

    return True


# =============================================================================
# MEDIUM PRIORITY: PART CONTENT RESTRICTIONS
# =============================================================================

def check_part_content(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate that part element only contains allowed children.
    Part can contain: partintro?, (appendix|chapter|article|preface|refentry|reference|glossary|bibliography|index|toc)+
    """
    if elem.tag != 'part':
        return True

    elem_id = elem.get('id', 'unknown')

    # Header elements allowed before partintro
    header_tags = {'beginpage', 'partinfo', 'title', 'subtitle', 'titleabbrev'}
    # Elements that are valid top-level children (per DTD)
    allowed_in_part = header_tags | {
        'partintro', 'appendix', 'chapter', 'article', 'preface',
        'refentry', 'reference', 'glossary', 'bibliography', 'index', 'toc', 'lot',
        # Note: anchor, indexterm should be inside partintro or chapters, not directly in part
    }

    for child in elem:
        if child.tag not in allowed_in_part:
            # Convert sect1 to chapter if found
            if child.tag == 'sect1':
                child.tag = 'chapter'
                issue = ValidationIssue(
                    element_tag='part',
                    element_id=elem_id,
                    issue_type='invalid_content',
                    description=f"<sect1> not allowed directly in <part>, converted to chapter",
                    severity='warning',
                    auto_fixed=True,
                    fix_description="Converted sect1 to chapter"
                )
                report.add_issue(issue)
            elif child.tag in {'para', 'simpara', 'formalpara', 'anchor', 'figure',
                               'table', 'mediaobject', 'sect1', 'indexterm'}:
                # These elements should be inside partintro, not directly in part
                # Wrap in partintro if no partintro exists
                partintro = elem.find('partintro')
                if partintro is None:
                    partintro = etree.Element('partintro')
                    # Find where to insert partintro (after header elements)
                    insert_idx = 0
                    for i, c in enumerate(elem):
                        if c.tag in header_tags:
                            insert_idx = i + 1
                        else:
                            break
                    elem.insert(insert_idx, partintro)
                # Move element to partintro
                elem.remove(child)
                partintro.append(child)
                issue = ValidationIssue(
                    element_tag='part',
                    element_id=elem_id,
                    issue_type='invalid_content',
                    description=f"<{child.tag}> not allowed directly in <part>, moved to partintro",
                    severity='warning',
                    auto_fixed=True,
                    fix_description=f"Moved {child.tag} to partintro"
                )
                report.add_issue(issue)
            else:
                issue = ValidationIssue(
                    element_tag='part',
                    element_id=elem_id,
                    issue_type='invalid_content',
                    description=f"<{child.tag}> not allowed inside <part>",
                    severity='error',
                    auto_fixed=False
                )
                report.add_issue(issue)

    return True


# =============================================================================
# LOW PRIORITY: INDEXTERM COMPLEX STRUCTURE VALIDATION
# =============================================================================

def check_indexterm_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate indexterm complex structure.
    Content model: (primary?, ((secondary, ((tertiary, (see|seealso+)?) | see | seealso+)?) | see | seealso+)?)

    Simplified rules:
    - If has primary, can have secondary (which can have tertiary)
    - see/seealso can appear at any level
    - class attribute affects requirements
    """
    if elem.tag != 'indexterm':
        return True

    elem_id = elem.get('id', 'unknown')
    idx_class = elem.get('class')

    # Check class attribute semantics
    if idx_class == 'startofrange':
        # Must have id, must NOT have startref
        if not elem.get('id'):
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<indexterm class='startofrange'> requires 'id' attribute",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)
        if elem.get('startref'):
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem_id,
                issue_type='invalid_attr',
                description=f"<indexterm class='startofrange'> should not have 'startref' attribute",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    elif idx_class == 'endofrange':
        # Must have startref (checked in referential integrity)
        # Should NOT have content (primary, secondary, etc.)
        if elem.find('primary') is not None or elem.find('see') is not None:
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem_id,
                issue_type='invalid_content',
                description=f"<indexterm class='endofrange'> should not have content elements",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # Check structure: secondary requires primary, tertiary requires secondary
    has_primary = elem.find('primary') is not None
    has_secondary = elem.find('secondary') is not None
    has_tertiary = elem.find('tertiary') is not None

    if has_secondary and not has_primary:
        # Add empty primary
        primary = etree.Element('primary')
        elem.insert(0, primary)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<secondary> requires <primary> in indexterm",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty primary"
        )
        report.add_issue(issue)

    if has_tertiary and not has_secondary:
        # Add empty secondary after primary
        secondary = etree.Element('secondary')
        primary = elem.find('primary')
        if primary is not None:
            idx = list(elem).index(primary) + 1
        else:
            idx = 0
        elem.insert(idx, secondary)
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='missing_required',
            description=f"<tertiary> requires <secondary> in indexterm",
            severity='error',
            auto_fixed=True,
            fix_description="Added empty secondary"
        )
        report.add_issue(issue)

    return True


# =============================================================================
# LOW PRIORITY: ATTRIBUTE SEMANTICS VALIDATION
# =============================================================================

def check_attribute_semantics(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate attribute semantics beyond just presence/absence.
    """
    tag = elem.tag
    elem_id = elem.get('id', 'unknown')

    # step/@performance validation
    if tag == 'step':
        performance = elem.get('performance')
        if performance and performance not in {'optional', 'required'}:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_attr_value',
                description=f"<step performance='{performance}'> invalid; must be 'optional' or 'required'",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # table/@frame validation
    if tag == 'table':
        frame = elem.get('frame')
        valid_frames = {'all', 'bottom', 'none', 'sides', 'top', 'topbot'}
        if frame and frame not in valid_frames:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_attr_value',
                description=f"<table frame='{frame}'> invalid; must be one of {valid_frames}",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # orderedlist/@numeration validation
    if tag == 'orderedlist':
        numeration = elem.get('numeration')
        valid_nums = {'arabic', 'loweralpha', 'lowerroman', 'upperalpha', 'upperroman'}
        if numeration and numeration not in valid_nums:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_attr_value',
                description=f"<orderedlist numeration='{numeration}'> invalid",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # simplelist/@type validation
    if tag == 'simplelist':
        list_type = elem.get('type')
        valid_types = {'inline', 'vert', 'horiz'}
        if list_type and list_type not in valid_types:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_attr_value',
                description=f"<simplelist type='{list_type}'> invalid; must be inline, vert, or horiz",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # bridgehead/@renderas validation
    if tag == 'bridgehead':
        renderas = elem.get('renderas')
        valid_render = {'sect1', 'sect2', 'sect3', 'sect4', 'sect5', 'other'}
        if renderas and renderas not in valid_render:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_attr_value',
                description=f"<bridgehead renderas='{renderas}'> invalid",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

    # imagedata/@align validation
    if tag == 'imagedata':
        align = elem.get('align')
        valid_aligns = {'left', 'right', 'center'}
        if align and align not in valid_aligns:
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='invalid_attr_value',
                description=f"<imagedata align='{align}'> invalid; must be left, right, or center",
                severity='warning',
                auto_fixed=False
            )
            report.add_issue(issue)

        # Check for fileref or entityref
        if not elem.get('fileref') and not elem.get('entityref'):
            issue = ValidationIssue(
                element_tag=tag,
                element_id=elem_id,
                issue_type='missing_required_attr',
                description=f"<imagedata> requires 'fileref' or 'entityref' attribute",
                severity='error',
                auto_fixed=False
            )
            report.add_issue(issue)

    return True


# =============================================================================
# LOW PRIORITY: BOOK/SET LEVEL VALIDATION
# =============================================================================

def check_book_structure(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate book-level structure requirements.
    """
    if elem.tag != 'book':
        return True

    elem_id = elem.get('id', 'unknown')

    # Book should have some content (chapters, parts, etc.)
    content_elements = {'dedication', 'toc', 'lot', 'glossary', 'bibliography',
                        'preface', 'chapter', 'reference', 'part', 'article',
                        'appendix', 'index', 'colophon'}

    has_content = any(elem.find(c) is not None for c in content_elements)

    if not has_content:
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem_id,
            issue_type='empty_element',
            description=f"<book> should have at least one content element (chapter, part, etc.)",
            severity='warning',
            auto_fixed=False
        )
        report.add_issue(issue)

    return True


def remove_duplicate_children(elem: etree.Element, child_tag: str, max_count: int,
                              report: ComplianceReport) -> None:
    """
    Remove duplicate children that violate DTD cardinality constraints.
    """
    children = [c for c in elem if c.tag == child_tag]
    if len(children) > max_count:
        for extra in children[max_count:]:
            elem.remove(extra)
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem.get('id', 'unknown'),
                issue_type='cardinality',
                description=f"<{elem.tag}> allows max {max_count} <{child_tag}> element(s)",
                severity='warning',
                auto_fixed=True,
                fix_description=f"Removed extra {child_tag} element"
            )
            report.add_issue(issue)


# =============================================================================
# CONTENT MODEL VALIDATION FOR splitContentFiles COMPATIBILITY
# =============================================================================

def check_dedication_content_model(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate dedication content model - only legalnotice.mix elements allowed.

    Per DTD: dedication = (risinfo?, (%sect.title.content;)?, (%legalnotice.mix;)+)

    legalnotice.mix includes:
    - glosslist, itemizedlist, orderedlist (lists)
    - caution, important, note, tip, warning (admonitions)
    - literallayout, programlisting, screen, synopsis, address (linespecific)
    - formalpara, para, simpara (paragraphs)
    - blockquote, indexterm, beginpage

    NOT allowed: anchor, figure, table, sidebar, mediaobject, sect1, etc.
    """
    if elem.tag != 'dedication':
        return True

    elem_id = elem.get('id', 'unknown')

    # Allowed elements in dedication (from LEGALNOTICE_MIX + title elements + risinfo)
    allowed_in_dedication = LEGALNOTICE_MIX | {
        'title', 'subtitle', 'titleabbrev',  # sect.title.content
        'risinfo',  # RIT-specific addition
    }

    # Elements that must be wrapped in para
    elements_needing_wrap = {'anchor', 'indexterm'}

    issues_found = 0

    for child in list(elem):
        child_tag = child.tag if isinstance(child.tag, str) else ''

        if child_tag and child_tag not in allowed_in_dedication:
            # Check if this is an element that can be wrapped in para
            if child_tag in elements_needing_wrap:
                # Wrap in para
                para = etree.Element('para')
                idx = list(elem).index(child)
                elem.remove(child)
                para.append(child)
                elem.insert(idx, para)

                issue = ValidationIssue(
                    element_tag='dedication',
                    element_id=elem_id,
                    issue_type='invalid_content_model',
                    description=f"<{child_tag}> not allowed directly in <dedication> (legalnotice.mix only)",
                    severity='error',
                    auto_fixed=True,
                    fix_description=f"Wrapped <{child_tag}> in <para>"
                )
                report.add_issue(issue)
                issues_found += 1
            elif child_tag == 'sect1':
                # sect1 is NOT allowed in dedication - cannot auto-fix
                issue = ValidationIssue(
                    element_tag='dedication',
                    element_id=elem_id,
                    issue_type='invalid_content_model',
                    description=f"<sect1> not allowed in <dedication> - dedication cannot have sections",
                    severity='error',
                    auto_fixed=False,
                    fix_description=None
                )
                report.add_issue(issue)
                issues_found += 1
            elif child_tag in {'figure', 'informalfigure', 'table', 'informaltable',
                               'sidebar', 'mediaobject', 'graphic', 'example',
                               'informalexample', 'procedure', 'bridgehead'}:
                # These elements are NOT in legalnotice.mix - cannot auto-fix
                issue = ValidationIssue(
                    element_tag='dedication',
                    element_id=elem_id,
                    issue_type='invalid_content_model',
                    description=f"<{child_tag}> not allowed in <dedication> (not in legalnotice.mix)",
                    severity='error',
                    auto_fixed=False,
                    fix_description=None
                )
                report.add_issue(issue)
                issues_found += 1

    return issues_found == 0


def check_bookcomponent_content_order(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate bookcomponent.content model for chapter/preface/appendix.

    Per DTD: bookcomponent.content =
        ((%divcomponent.mix;)+, (sect1* | ...)) | (sect1+ | ...)

    CRITICAL RULE: Once sect1 elements start, NO block content can appear
    after them at the same level. This causes splitContentFiles failures.
    """
    if elem.tag not in {'chapter', 'preface', 'appendix'}:
        return True

    elem_id = elem.get('id', 'unknown')

    # Section elements
    section_tags = {'sect1', 'simplesect', 'section', 'risempty'}

    # Metadata elements to skip
    meta_tags = {'title', 'subtitle', 'titleabbrev', 'chapterinfo', 'prefaceinfo',
                 'appendixinfo', 'beginpage', 'tocchap'}

    # Navigation elements allowed after sections
    nav_tags = {'toc', 'lot', 'index', 'glossary', 'bibliography'}

    children = list(elem)
    first_sect_idx = None

    # Find the first section element
    for i, child in enumerate(children):
        child_tag = child.tag if isinstance(child.tag, str) else ''
        if child_tag in section_tags:
            first_sect_idx = i
            break

    if first_sect_idx is None:
        return True  # No sections, no issue

    # Check for block content after first section
    issues_found = 0

    for i in range(first_sect_idx + 1, len(children)):
        child = children[i]
        child_tag = child.tag if isinstance(child.tag, str) else ''

        # Skip allowed elements
        if child_tag in section_tags or child_tag in nav_tags or child_tag in meta_tags:
            continue

        # This is block content after sections - invalid!
        if child_tag in DIVCOMPONENT_MIX:
            issue = ValidationIssue(
                element_tag=elem.tag,
                element_id=elem_id,
                issue_type='invalid_content_model',
                description=f"<{child_tag}> appears after <sect1> in <{elem.tag}> - block content must come BEFORE sections",
                severity='error',
                auto_fixed=False,
                fix_description=None
            )
            report.add_issue(issue)
            issues_found += 1

    return issues_found == 0


def check_sect1_id_required(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate that all sect1 elements have @id attribute with correct format.

    Sect1 @id is REQUIRED for:
    1. Chunking by RISChunker.xsl (creates sect1.{isbn}.{id}.xml files)
    2. R2 Library navigation (parses linkend to find target file)
    3. TOC generation

    CRITICAL: Sect1 IDs must be EXACTLY {prefix}####s#### (e.g., ch0001s0001).
    IDs like ch0005s0001s10002 are INVALID - they look like sect2 hierarchical IDs.
    This function auto-fixes malformed sect1 IDs by regenerating them.
    """
    if elem.tag != 'sect1':
        return True

    # Find root and parent for ID generation
    root = elem
    while root.getparent() is not None:
        root = root.getparent()

    parent = elem.getparent()
    parent_id = parent.get('id', '') if parent is not None else ''

    if not elem.get('id'):
        # Generate a new ID for missing sect1 @id
        if not parent_id:
            # Extract chapter prefix from context
            id_match = re.match(r'^([a-z]{2}\d{4})', elem.get('id', ''))
            parent_id = id_match.group(1) if id_match else 'ch0001'

        existing_ids = {e.get('id') for e in root.iter() if e.get('id')}
        new_id = next_available_sect1_id(parent_id, existing_ids)
        elem.set('id', new_id)

        issue = ValidationIssue(
            element_tag='sect1',
            element_id=new_id,
            issue_type='missing_required_attr',
            description=f"<sect1> requires 'id' attribute for chunking (parent: {parent_id})",
            severity='error',
            auto_fixed=True,
            fix_description=f"Generated ID: {new_id}"
        )
        report.add_issue(issue)
        return True

    # Validate ID format - must be EXACTLY {prefix}####s#### (11 characters)
    sect1_id = elem.get('id')

    # ID should follow EXACT pattern: {prefix}{4-digit}s{4-digit} with NO additional segments
    # The $ anchor ensures we match the ENTIRE ID, not just a prefix
    # This catches malformed IDs like ch0005s0001s10002 which have extra hierarchical segments
    valid_pattern = re.compile(r'^[a-z]{2}\d{4}s\d{4}$')
    if not valid_pattern.match(sect1_id):
        # This is a malformed sect1 ID - regenerate it
        old_id = sect1_id

        # Extract chapter prefix from the malformed ID
        prefix_match = re.match(r'^([a-z]{2}\d{4})', sect1_id)
        chapter_prefix = prefix_match.group(1) if prefix_match else (parent_id if parent_id else 'ch0001')

        # Collect existing IDs
        existing_ids = {e.get('id') for e in root.iter() if e.get('id')}

        # Generate new compliant sect1 ID
        new_id = next_available_sect1_id(chapter_prefix, existing_ids)
        elem.set('id', new_id)

        # Update child element IDs that used the old section ID as a prefix
        child_updates = _update_child_element_ids(elem, old_id, new_id)

        # Update linkend references throughout document
        linkend_updates = _update_linkend_references(root, old_id, new_id)

        fix_desc = f"Regenerated sect1 ID from malformed '{old_id}' to '{new_id}'"
        if child_updates > 0:
            fix_desc += f", updated {child_updates} child IDs"
        if linkend_updates > 0:
            fix_desc += f", updated {linkend_updates} linkend refs"

        issue = ValidationIssue(
            element_tag='sect1',
            element_id=new_id,
            issue_type='invalid_id_format',
            description=f"sect1 @id '{old_id}' was malformed (must be EXACTLY {{prefix}}####s####)",
            severity='warning',
            auto_fixed=True,
            fix_description=fix_desc
        )
        report.add_issue(issue)
        logger.info(f"Fixed malformed sect1 ID: '{old_id}' -> '{new_id}'")

    return True


def check_toc_content_model(elem: etree.Element, report: ComplianceReport) -> bool:
    """
    Validate TOC content model - only toc* elements allowed.

    Per DTD: toc = (beginpage?, (%bookcomponent.title.content;)?,
                    tocfront*, (tocpart | tocchap)*, tocback*)

    NOT allowed: sect1, para, figure, etc.
    """
    if elem.tag != 'toc':
        return True

    elem_id = elem.get('id', 'unknown')

    # Allowed elements in toc
    allowed_in_toc = {
        'beginpage', 'title', 'subtitle', 'titleabbrev',
        'tocfront', 'tocpart', 'tocchap', 'tocback',
        'tocentry', 'toclevel1', 'toclevel2', 'toclevel3', 'toclevel4', 'toclevel5'
    }

    issues_found = 0

    for child in list(elem):
        child_tag = child.tag if isinstance(child.tag, str) else ''

        if child_tag and child_tag not in allowed_in_toc:
            issue = ValidationIssue(
                element_tag='toc',
                element_id=elem_id,
                issue_type='invalid_content_model',
                description=f"<{child_tag}> not allowed in <toc> - use toc* elements (tocchap, tocentry, etc.)",
                severity='error',
                auto_fixed=False,
                fix_description=None
            )
            report.add_issue(issue)
            issues_found += 1

    return issues_found == 0


def check_toc_linkend_validity(elem: etree.Element, all_ids: set, report: ComplianceReport) -> bool:
    """
    Validate that all TOC linkend attributes point to valid IDs.

    This is called during referential integrity check.
    """
    toc_elements = {'tocentry', 'tocfront', 'tocback'}

    if elem.tag not in toc_elements:
        return True

    linkend = elem.get('linkend')
    if not linkend:
        return True  # linkend is optional

    if linkend not in all_ids:
        issue = ValidationIssue(
            element_tag=elem.tag,
            element_id=elem.get('id', 'unknown'),
            issue_type='broken_reference',
            description=f"<{elem.tag} linkend='{linkend}'> references non-existent ID",
            severity='error',
            auto_fixed=False,
            fix_description=None
        )
        report.add_issue(issue)
        return False

    return True


# =============================================================================
# MAIN VALIDATION FUNCTION
# =============================================================================

def validate_and_fix_dtd_compliance(root_elem: etree.Element) -> ComplianceReport:
    """
    Comprehensive DTD compliance validation and auto-fix.

    Args:
        root_elem: Root element of the XML tree

    Returns:
        ComplianceReport with all issues found and fix status
    """
    report = ComplianceReport()

    logger.info("Starting comprehensive DTD compliance check...")

    # Iterate through all elements
    for elem in root_elem.iter():
        # Skip text nodes
        if not isinstance(elem.tag, str):
            continue

        # 1. Check element ordering (beginpage before title, etc.)
        check_element_ordering(elem, report)

        # 2. Check for empty elements that require content
        check_empty_element(elem, report)

        # 3. Check required attributes
        check_required_attributes(elem, report)

        # 4. Check specific structural requirements
        check_table_structure(elem, report)
        check_tgroup_ordering(elem, report)
        check_list_structure(elem, report)
        check_segmented_list_structure(elem, report)
        check_callout_list_structure(elem, report)
        check_figure_structure(elem, report)
        check_bibliography_structure(elem, report)
        check_glossary_structure(elem, report)
        check_mediaobject_structure(elem, report)
        check_textobject_structure(elem, report)
        check_imageobject_structure(elem, report)
        check_invalid_nesting(elem, report)

        # 5. Check complex element structures
        check_formalpara_structure(elem, report)
        check_procedure_structure(elem, report)
        check_substeps_structure(elem, report)
        check_refentry_structure(elem, report)
        check_refnamediv_structure(elem, report)
        check_refsect_structure(elem, report)
        check_info_element_structure(elem, report)
        check_abstract_structure(elem, report)
        check_qandaentry_structure(elem, report)
        check_answer_structure(elem, report)

        # 5b. Check content models for splitContentFiles compatibility
        check_dedication_content_model(elem, report)
        check_bookcomponent_content_order(elem, report)
        check_toc_content_model(elem, report)

        # 6. Check nesting constraints - MUST run before check_sect1_id_required
        # so that sect2->sect1 conversions happen first, then IDs are validated
        check_section_nesting(elem, report)

        # 6b. Validate sect1 IDs - runs AFTER check_section_nesting to catch
        # any sect1 elements that were converted from sect2 and may need ID fixes
        check_sect1_id_required(elem, report)

        # 7. Check cardinality constraints
        # Only one beginpage per container
        if elem.tag in {'chapter', 'appendix', 'preface', 'sect1', 'sect2',
                        'sect3', 'sect4', 'sect5', 'section'}:
            remove_duplicate_children(elem, 'beginpage', 1, report)

        # Only one title per container
        if elem.tag in {'chapter', 'appendix', 'preface', 'sect1', 'sect2',
                        'sect3', 'sect4', 'sect5', 'section', 'figure',
                        'table', 'example', 'sidebar'}:
            remove_duplicate_children(elem, 'title', 1, report)

        # 8. Check content restrictions by parent context
        check_content_restrictions(elem, report)

        # 9. Check choice group constraints
        check_choice_group_constraints(elem, report)

        # 10. Check part content restrictions
        check_part_content(elem, report)

        # 11. Check indexterm structure
        check_indexterm_structure(elem, report)

        # 12. Check attribute semantics
        check_attribute_semantics(elem, report)

    # 13. Referential integrity (needs full tree traversal)
    check_referential_integrity(root_elem, report)

    # 14. Book-level structure validation
    check_book_structure(root_elem, report)

    logger.info(report.get_summary())

    return report


def generate_compliance_report_xlsx(report: ComplianceReport, output_path: str) -> None:
    """
    Generate an Excel report of all compliance issues.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "DTD Compliance Issues"

        # Header row
        headers = ['Element', 'ID', 'Issue Type', 'Description', 'Severity', 'Auto-Fixed', 'Fix Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data rows
        for row_idx, issue in enumerate(report.issues, 2):
            ws.cell(row=row_idx, column=1, value=issue.element_tag)
            ws.cell(row=row_idx, column=2, value=issue.element_id)
            ws.cell(row=row_idx, column=3, value=issue.issue_type)
            ws.cell(row=row_idx, column=4, value=issue.description)
            ws.cell(row=row_idx, column=5, value=issue.severity)
            ws.cell(row=row_idx, column=6, value="Yes" if issue.auto_fixed else "No")
            ws.cell(row=row_idx, column=7, value=issue.fix_description or "")

            # Color unfixed errors red
            if not issue.auto_fixed and issue.severity == 'error':
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )

        # Summary row
        summary_row = len(report.issues) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f"Total Issues: {len(report.issues)}")
        ws.cell(row=summary_row, column=3, value=f"Auto-Fixed: {report.fixed_count}")
        ws.cell(row=summary_row, column=4, value=f"Require Attention: {report.unfixed_count}")

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30

        wb.save(output_path)
        logger.info(f"Compliance report saved to {output_path}")

    except ImportError:
        logger.warning("openpyxl not available, skipping Excel report generation")
    except Exception as e:
        logger.error(f"Failed to generate compliance report: {e}")
