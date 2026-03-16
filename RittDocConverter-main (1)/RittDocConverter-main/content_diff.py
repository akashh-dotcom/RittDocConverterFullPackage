#!/usr/bin/env python3
"""
Content Diff Tool for RittDoc Converter

This module compares the text content between the original EPUB and the converted
XML package (rittdoc zip) to identify content differences and potential content loss.

Outputs:
- contentdiff.xlsx: Detailed comparison with line-by-line differences
- Adds entry to validation_report.xlsx if content loss is detected
"""

import logging
import re
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import OrderedDict
from difflib import SequenceMatcher, unified_diff

from bs4 import BeautifulSoup, NavigableString, Tag
from lxml import etree
import ebooklib
from ebooklib import epub

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not available. Excel output will be skipped.")


class ContentDiffResult:
    """Container for content diff results."""

    def __init__(self):
        self.epub_chapters: Dict[str, str] = OrderedDict()  # chapter_id -> text content
        self.xml_chapters: Dict[str, str] = OrderedDict()   # chapter_id -> text content
        self.differences: List[Dict] = []  # List of difference records
        self.similarity_scores: Dict[str, float] = {}  # chapter_id -> similarity %
        self.total_epub_chars: int = 0
        self.total_xml_chars: int = 0
        self.content_loss_detected: bool = False
        self.content_loss_percentage: float = 0.0
        self.missing_chapters: List[str] = []
        self.extra_chapters: List[str] = []


def extract_text_from_html(html_content: bytes, normalize: bool = True) -> str:
    """
    Extract plain text content from HTML/XHTML content.

    Args:
        html_content: Raw HTML/XHTML bytes
        normalize: If True, normalize whitespace

    Returns:
        Extracted text content
    """
    try:
        soup = BeautifulSoup(html_content, 'html.parser')

        # Remove script and style elements
        for element in soup(['script', 'style', 'head', 'meta', 'link']):
            element.decompose()

        # Get text
        text = soup.get_text(separator=' ')

        if normalize:
            # Normalize whitespace
            text = re.sub(r'\s+', ' ', text).strip()

        return text
    except Exception as e:
        logger.error(f"Error extracting text from HTML: {e}")
        return ""


def extract_text_from_xml(xml_content: bytes, normalize: bool = True) -> str:
    """
    Extract plain text content from DocBook XML content.

    Args:
        xml_content: Raw XML bytes
        normalize: If True, normalize whitespace

    Returns:
        Extracted text content
    """
    try:
        # Parse XML
        parser = etree.XMLParser(recover=True, remove_blank_text=True)
        root = etree.fromstring(xml_content, parser=parser)

        # Extract all text content
        text_parts = []
        for elem in root.iter():
            if elem.text:
                text_parts.append(elem.text)
            if elem.tail:
                text_parts.append(elem.tail)

        text = ' '.join(text_parts)

        if normalize:
            # Normalize whitespace
            text = re.sub(r'\s+', ' ', text).strip()

        return text
    except Exception as e:
        logger.error(f"Error extracting text from XML: {e}")
        return ""


def extract_epub_content(epub_path: Path) -> Dict[str, str]:
    """
    Extract text content from all chapters in an EPUB file.

    Args:
        epub_path: Path to EPUB file

    Returns:
        Dict mapping chapter IDs to text content
    """
    chapters = OrderedDict()

    try:
        book = epub.read_epub(str(epub_path))

        # Get spine items
        for idx, (item_id, _) in enumerate(book.spine):
            item = book.get_item_with_id(item_id)
            if item and item.get_type() == ebooklib.ITEM_DOCUMENT:
                chapter_id = f"ch{idx+1:04d}"
                content = item.get_content()
                text = extract_text_from_html(content)
                chapters[chapter_id] = text
                logger.debug(f"Extracted {len(text)} chars from EPUB {chapter_id}")
    except Exception as e:
        logger.error(f"Error reading EPUB {epub_path}: {e}")

    return chapters


def _get_entity_order_from_book_xml(zf: zipfile.ZipFile) -> List[str]:
    """
    Read Book.XML from a rittdoc zip and return entity references in document order.

    Entity references like &ch0001;, &pt0001;, &ap0001; etc. define the spine
    order of fragment files within the package.

    Args:
        zf: Open ZipFile object

    Returns:
        List of entity names in document order (e.g., ['pr0001', 'ch0001', 'pt0001', ...])
    """
    for name in zf.namelist():
        if Path(name).name.lower() == 'book.xml':
            try:
                content = zf.read(name).decode('utf-8')
                # Find all entity references (e.g., &ch0001;) in order of appearance
                return re.findall(r'&([a-z]{2}\d{4});', content)
            except Exception:
                pass
            break
    return []


def extract_rittdoc_content(rittdoc_path: Path) -> Dict[str, str]:
    """
    Extract text content from ALL XML fragment files in a rittdoc zip package.

    Includes all element types: chapters (ch), appendices (ap), prefaces (pr),
    bibliographies (bi), dedications (dd), glossaries (gl), indices (in),
    parts (pt), subparts (sp), and any other 2-letter-prefix fragment files.

    Files are returned in Book.XML entity reference order when available,
    with any remaining files appended in alphabetical order.

    Args:
        rittdoc_path: Path to rittdoc zip file

    Returns:
        Dict mapping element IDs to text content (e.g., ch0001, ap0001, pt0001)
    """
    chapters = OrderedDict()

    try:
        with zipfile.ZipFile(rittdoc_path, 'r') as zf:
            # Read Book.XML to determine the entity reference order
            entity_order = _get_entity_order_from_book_xml(zf)

            # Find ALL fragment XML files matching the element-type pattern
            # (2-letter prefix + 4-digit number, e.g., ch0001.xml, pt0001.xml, ap0001.xml)
            xml_files = [f for f in zf.namelist()
                        if f.endswith('.xml') and re.match(r'[a-z]{2}\d{4}\.xml', Path(f).name)]

            # Build lookup by stem (e.g., "ch0001" -> "path/ch0001.xml")
            xml_by_stem = {}
            for xml_file in xml_files:
                stem = Path(xml_file).stem
                xml_by_stem[stem] = xml_file

            # Order files: Book.XML entity references first, then remaining alphabetically
            ordered_stems = []
            seen = set()
            for entity in entity_order:
                if entity in xml_by_stem and entity not in seen:
                    ordered_stems.append(entity)
                    seen.add(entity)
            for stem in sorted(xml_by_stem.keys()):
                if stem not in seen:
                    ordered_stems.append(stem)
                    seen.add(stem)

            for stem in ordered_stems:
                xml_file = xml_by_stem[stem]
                try:
                    content = zf.read(xml_file)
                    text = extract_text_from_xml(content)
                    chapters[stem] = text
                    logger.debug(f"Extracted {len(text)} chars from XML {stem}")
                except Exception as e:
                    logger.error(f"Error reading {xml_file}: {e}")
    except Exception as e:
        logger.error(f"Error reading rittdoc {rittdoc_path}: {e}")

    return chapters


def compute_similarity(text1: str, text2: str, quick_mode: bool = True) -> float:
    """
    Compute similarity ratio between two texts.

    Args:
        text1: First text
        text2: Second text
        quick_mode: If True, use faster approximate comparison for large texts

    Returns:
        Similarity ratio (0.0 to 1.0)
    """
    if not text1 and not text2:
        return 1.0
    if not text1 or not text2:
        return 0.0

    # For identical texts, return 1.0 immediately
    if text1 == text2:
        return 1.0

    # Quick length-based pre-check
    len1, len2 = len(text1), len(text2)
    if len1 > 0 and len2 > 0:
        length_ratio = min(len1, len2) / max(len1, len2)
        # If lengths differ by more than 50%, texts are quite different
        if length_ratio < 0.5:
            return length_ratio * 0.5  # Rough estimate

    # For large texts, use faster word-based comparison
    if quick_mode and (len1 > 10000 or len2 > 10000):
        return _compute_similarity_fast(text1, text2)

    # For smaller texts, use standard SequenceMatcher
    # Use autojunk=False for more accurate results on shorter texts
    return SequenceMatcher(None, text1, text2, autojunk=False).ratio()


def _compute_similarity_fast(text1: str, text2: str) -> float:
    """
    Fast approximate similarity using word set comparison (Jaccard similarity).

    This is O(n) instead of O(n²) for SequenceMatcher.

    Args:
        text1: First text
        text2: Second text

    Returns:
        Approximate similarity ratio (0.0 to 1.0)
    """
    # Split into words and create sets
    words1 = set(text1.lower().split())
    words2 = set(text2.lower().split())

    if not words1 and not words2:
        return 1.0
    if not words1 or not words2:
        return 0.0

    # Jaccard similarity: intersection / union
    intersection = len(words1 & words2)
    union = len(words1 | words2)

    if union == 0:
        return 1.0

    jaccard = intersection / union

    # Also consider word count similarity for a more balanced score
    count1, count2 = len(text1.split()), len(text2.split())
    count_ratio = min(count1, count2) / max(count1, count2) if max(count1, count2) > 0 else 1.0

    # Combine Jaccard and count ratio (weighted average)
    return 0.7 * jaccard + 0.3 * count_ratio


def find_differences(text1: str, text2: str, context_chars: int = 50, max_differences: int = 100) -> List[Dict]:
    """
    Find specific differences between two texts.

    Args:
        text1: Original text (EPUB)
        text2: Converted text (XML)
        context_chars: Number of context characters to include
        max_differences: Maximum number of differences to return (for performance)

    Returns:
        List of difference records
    """
    differences = []

    # For very large texts, use a sampling approach
    words1 = text1.split()
    words2 = text2.split()

    # If texts are very large, sample for performance
    MAX_WORDS = 5000
    sampled = False
    if len(words1) > MAX_WORDS or len(words2) > MAX_WORDS:
        # Sample beginning, middle, and end
        sampled = True
        sample_size = MAX_WORDS // 3

        if len(words1) > MAX_WORDS:
            mid_start1 = len(words1) // 2 - sample_size // 2
            words1_sampled = words1[:sample_size] + words1[mid_start1:mid_start1 + sample_size] + words1[-sample_size:]
        else:
            words1_sampled = words1

        if len(words2) > MAX_WORDS:
            mid_start2 = len(words2) // 2 - sample_size // 2
            words2_sampled = words2[:sample_size] + words2[mid_start2:mid_start2 + sample_size] + words2[-sample_size:]
        else:
            words2_sampled = words2

        words1, words2 = words1_sampled, words2_sampled

    # Use autojunk=True for faster processing on large texts
    matcher = SequenceMatcher(None, words1, words2, autojunk=True)

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            continue

        # Stop if we've found enough differences
        if len(differences) >= max_differences:
            break

        # Get context
        context_before = ' '.join(words1[max(0, i1-5):i1])
        context_after = ' '.join(words1[i2:min(len(words1), i2+5)])

        if tag == 'delete':
            # Content in EPUB but not in XML (potential loss)
            deleted_text = ' '.join(words1[i1:i2])
            differences.append({
                'type': 'deleted',
                'epub_text': deleted_text[:500],  # Limit text length
                'xml_text': '',
                'context_before': context_before,
                'context_after': context_after,
                'position': i1
            })
        elif tag == 'insert':
            # Content in XML but not in EPUB (additions)
            inserted_text = ' '.join(words2[j1:j2])
            differences.append({
                'type': 'inserted',
                'epub_text': '',
                'xml_text': inserted_text[:500],  # Limit text length
                'context_before': context_before,
                'context_after': context_after,
                'position': i1
            })
        elif tag == 'replace':
            # Content changed
            original_text = ' '.join(words1[i1:i2])
            replaced_text = ' '.join(words2[j1:j2])
            differences.append({
                'type': 'replaced',
                'epub_text': original_text[:500],  # Limit text length
                'xml_text': replaced_text[:500],  # Limit text length
                'context_before': context_before,
                'context_after': context_after,
                'position': i1
            })

    if sampled and len(differences) > 0:
        differences[0]['note'] = '(sampled comparison - full text was too large)'

    return differences


def _match_chapters_by_content(
    unmatched_epub: Dict[str, str],
    unmatched_xml: Dict[str, str],
    min_similarity: float = 0.3
) -> Dict[str, str]:
    """
    Match unmatched EPUB chapters to unmatched rittdoc chapters by content similarity.

    EPUB extraction assigns sequential ch#### IDs to ALL spine items, but
    rittdoc uses element-type-specific prefixes (ch, ap, pr, pt, dd, gl, in, sp).
    This function pairs them up based on which texts are most similar.

    Args:
        unmatched_epub: Dict of unmatched EPUB chapter IDs to text content
        unmatched_xml: Dict of unmatched rittdoc chapter IDs to text content
        min_similarity: Minimum similarity score to accept a match (0.0 to 1.0)

    Returns:
        Dict mapping EPUB chapter IDs to rittdoc chapter IDs (e.g., ch0018 -> pt0001)
    """
    if not unmatched_epub or not unmatched_xml:
        return {}

    # Compute similarity matrix (only for non-empty texts)
    scores = []  # (score, epub_id, xml_id)
    for epub_id, epub_text in unmatched_epub.items():
        if not epub_text:
            continue
        for xml_id, xml_text in unmatched_xml.items():
            if not xml_text:
                continue
            score = compute_similarity(epub_text, xml_text, quick_mode=True)
            if score >= min_similarity:
                scores.append((score, epub_id, xml_id))

    # Greedy matching: best scores first, each ID used at most once
    scores.sort(reverse=True)
    mapping = {}
    used_xml = set()
    for score, epub_id, xml_id in scores:
        if epub_id in mapping or xml_id in used_xml:
            continue
        mapping[epub_id] = xml_id
        used_xml.add(xml_id)

    return mapping


def compare_content(epub_path: Path, rittdoc_path: Path, skip_detailed_diff: bool = False) -> ContentDiffResult:
    """
    Compare content between EPUB and rittdoc package.

    Args:
        epub_path: Path to original EPUB file
        rittdoc_path: Path to converted rittdoc zip file
        skip_detailed_diff: If True, skip detailed diff for performance (just compare totals)

    Returns:
        ContentDiffResult with comparison results
    """
    import time
    result = ContentDiffResult()

    start_time = time.time()
    logger.info(f"Extracting content from EPUB: {epub_path}")
    result.epub_chapters = extract_epub_content(epub_path)
    logger.info(f"  Extracted {len(result.epub_chapters)} chapters from EPUB in {time.time() - start_time:.1f}s")

    start_time = time.time()
    logger.info(f"Extracting content from rittdoc: {rittdoc_path}")
    result.xml_chapters = extract_rittdoc_content(rittdoc_path)
    logger.info(f"  Extracted {len(result.xml_chapters)} chapters from rittdoc in {time.time() - start_time:.1f}s")

    # Calculate total characters
    result.total_epub_chars = sum(len(text) for text in result.epub_chapters.values())
    result.total_xml_chars = sum(len(text) for text in result.xml_chapters.values())

    logger.info(f"  EPUB total: {result.total_epub_chars:,} chars, XML total: {result.total_xml_chars:,} chars")

    # Smart matching: EPUB uses sequential ch#### IDs for ALL spine items,
    # but rittdoc uses element-type-specific prefixes (ch, ap, pr, pt, etc.).
    # First match by direct ID (ch0001-chNNNN will match directly),
    # then match remaining items by content similarity.
    epub_ids = set(result.epub_chapters.keys())
    xml_ids = set(result.xml_chapters.keys())

    direct_matches = epub_ids & xml_ids
    unmatched_epub_ids = sorted(epub_ids - direct_matches)
    unmatched_xml_ids = sorted(xml_ids - direct_matches)

    if unmatched_epub_ids and unmatched_xml_ids:
        # Try to pair unmatched EPUB items with unmatched rittdoc items by content
        id_mapping = _match_chapters_by_content(
            {eid: result.epub_chapters[eid] for eid in unmatched_epub_ids},
            {xid: result.xml_chapters[xid] for xid in unmatched_xml_ids}
        )

        if id_mapping:
            mapped_pairs = [f"{eid}\u2192{xid}" for eid, xid in sorted(id_mapping.items())]
            logger.info(f"  Matched {len(id_mapping)} EPUB items to rittdoc by content: {', '.join(mapped_pairs)}")

            # Re-key EPUB chapters: replace sequential ch#### IDs with rittdoc IDs
            rekeyed_epub = OrderedDict()
            for epub_id, epub_text in result.epub_chapters.items():
                new_id = id_mapping.get(epub_id, epub_id)
                rekeyed_epub[new_id] = epub_text
            result.epub_chapters = rekeyed_epub

    # Recalculate after re-keying
    epub_ids = set(result.epub_chapters.keys())
    xml_ids = set(result.xml_chapters.keys())
    result.missing_chapters = sorted(epub_ids - xml_ids)
    result.extra_chapters = sorted(xml_ids - epub_ids)

    # Compare each chapter
    all_chapters = sorted(epub_ids | xml_ids)
    total_similarity = 0.0
    compared_count = 0

    start_time = time.time()
    logger.info(f"Comparing {len(all_chapters)} chapters...")

    for idx, chapter_id in enumerate(all_chapters):
        epub_text = result.epub_chapters.get(chapter_id, '')
        xml_text = result.xml_chapters.get(chapter_id, '')

        # Progress logging for large comparisons
        if (idx + 1) % 10 == 0 or idx == len(all_chapters) - 1:
            elapsed = time.time() - start_time
            logger.info(f"  Processing chapter {idx + 1}/{len(all_chapters)} ({elapsed:.1f}s elapsed)")

        # Compute similarity (using fast mode for large texts)
        similarity = compute_similarity(epub_text, xml_text, quick_mode=True)
        result.similarity_scores[chapter_id] = similarity

        if epub_text or xml_text:
            total_similarity += similarity
            compared_count += 1

        # Find differences if similarity is not perfect and not skipping detailed diff
        if not skip_detailed_diff and similarity < 0.99:
            # Only do detailed diff for significantly different chapters
            chapter_diffs = find_differences(epub_text, xml_text, max_differences=50)
            for diff in chapter_diffs:
                diff['chapter_id'] = chapter_id
                result.differences.append(diff)

            # Limit total differences to prevent memory issues
            if len(result.differences) > 500:
                logger.warning("Too many differences found, stopping detailed comparison")
                skip_detailed_diff = True

    logger.info(f"  Comparison completed in {time.time() - start_time:.1f}s")

    # Calculate overall similarity and content loss
    if compared_count > 0:
        avg_similarity = total_similarity / compared_count
    else:
        avg_similarity = 1.0

    # Calculate content loss percentage
    if result.total_epub_chars > 0:
        char_diff = result.total_epub_chars - result.total_xml_chars
        if char_diff > 0:
            result.content_loss_percentage = (char_diff / result.total_epub_chars) * 100
            # Consider it significant loss if more than 1% or more than 100 chars
            result.content_loss_detected = (result.content_loss_percentage > 1.0 or char_diff > 100)

    # Also flag as loss if there are deleted content items
    deleted_count = sum(1 for d in result.differences if d['type'] == 'deleted')
    if deleted_count > 10:  # More than 10 deleted items
        result.content_loss_detected = True

    logger.info(f"Comparison complete: {len(result.differences)} differences found")
    logger.info(f"EPUB chars: {result.total_epub_chars:,}, XML chars: {result.total_xml_chars:,}")
    logger.info(f"Content loss: {result.content_loss_percentage:.2f}%")

    return result


def write_contentdiff_excel(result: ContentDiffResult, output_path: Path) -> bool:
    """
    Write content diff results to Excel file.

    Args:
        result: ContentDiffResult from comparison
        output_path: Path to output Excel file

    Returns:
        True if successful, False otherwise
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl not available. Cannot create Excel output.")
        return False

    try:
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Summary data
        summary_data = [
            ('EPUB Total Characters', result.total_epub_chars),
            ('XML Total Characters', result.total_xml_chars),
            ('Character Difference', result.total_epub_chars - result.total_xml_chars),
            ('Content Loss Percentage', f"{result.content_loss_percentage:.2f}%"),
            ('Content Loss Detected', 'YES' if result.content_loss_detected else 'NO'),
            ('Total Differences Found', len(result.differences)),
            ('Missing Chapters', ', '.join(result.missing_chapters) if result.missing_chapters else 'None'),
            ('Extra Chapters', ', '.join(result.extra_chapters) if result.extra_chapters else 'None'),
            ('EPUB Chapter Count', len(result.epub_chapters)),
            ('XML Chapter Count', len(result.xml_chapters)),
        ]

        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1, value=metric)
            ws_summary.cell(row=row, column=2, value=str(value))

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 50

        # Chapter Similarity sheet
        ws_similarity = wb.create_sheet("Chapter Similarity")

        headers = ['Chapter ID', 'EPUB Chars', 'XML Chars', 'Similarity %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws_similarity.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for chapter_id in sorted(set(result.epub_chapters.keys()) | set(result.xml_chapters.keys())):
            epub_chars = len(result.epub_chapters.get(chapter_id, ''))
            xml_chars = len(result.xml_chapters.get(chapter_id, ''))
            similarity = result.similarity_scores.get(chapter_id, 0.0) * 100

            if chapter_id in result.missing_chapters:
                status = 'MISSING'
            elif chapter_id in result.extra_chapters:
                status = 'EXTRA'
            elif similarity >= 99:
                status = 'OK'
            elif similarity >= 90:
                status = 'MINOR_DIFF'
            else:
                status = 'SIGNIFICANT_DIFF'

            ws_similarity.cell(row=row, column=1, value=chapter_id)
            ws_similarity.cell(row=row, column=2, value=epub_chars)
            ws_similarity.cell(row=row, column=3, value=xml_chars)
            ws_similarity.cell(row=row, column=4, value=f"{similarity:.1f}%")
            status_cell = ws_similarity.cell(row=row, column=5, value=status)

            # Color code status
            if status == 'MISSING' or status == 'SIGNIFICANT_DIFF':
                status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif status == 'MINOR_DIFF':
                status_cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
            elif status == 'OK':
                status_cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")

            row += 1

        # Adjust column widths
        for col in range(1, 6):
            ws_similarity.column_dimensions[get_column_letter(col)].width = 15

        # Differences sheet
        ws_diffs = wb.create_sheet("Differences")

        headers = ['Chapter', 'Type', 'EPUB Text', 'XML Text', 'Context Before', 'Context After']
        for col, header in enumerate(headers, 1):
            cell = ws_diffs.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for diff in result.differences[:1000]:  # Limit to first 1000 differences
            ws_diffs.cell(row=row, column=1, value=diff.get('chapter_id', ''))

            diff_type = diff.get('type', '')
            type_cell = ws_diffs.cell(row=row, column=2, value=diff_type)

            # Color code by type
            if diff_type == 'deleted':
                type_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif diff_type == 'inserted':
                type_cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
            elif diff_type == 'replaced':
                type_cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")

            # Truncate long texts
            epub_text = diff.get('epub_text', '')[:500]
            xml_text = diff.get('xml_text', '')[:500]
            context_before = diff.get('context_before', '')[:100]
            context_after = diff.get('context_after', '')[:100]

            ws_diffs.cell(row=row, column=3, value=epub_text)
            ws_diffs.cell(row=row, column=4, value=xml_text)
            ws_diffs.cell(row=row, column=5, value=context_before)
            ws_diffs.cell(row=row, column=6, value=context_after)

            row += 1

        # Adjust column widths
        ws_diffs.column_dimensions['A'].width = 10
        ws_diffs.column_dimensions['B'].width = 12
        ws_diffs.column_dimensions['C'].width = 60
        ws_diffs.column_dimensions['D'].width = 60
        ws_diffs.column_dimensions['E'].width = 30
        ws_diffs.column_dimensions['F'].width = 30

        # Save workbook
        wb.save(output_path)
        logger.info(f"Saved content diff to {output_path}")
        return True

    except Exception as e:
        logger.error(f"Error writing Excel file: {e}")
        return False


def add_to_validation_report(validation_report_path: Path, result: ContentDiffResult,
                             epub_name: str) -> bool:
    """
    Add content loss entry to validation report Excel file.

    Args:
        validation_report_path: Path to validation_report.xlsx
        result: ContentDiffResult from comparison
        epub_name: Name of the EPUB file

    Returns:
        True if successful, False otherwise
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl not available. Cannot update validation report.")
        return False

    if not result.content_loss_detected:
        logger.info("No significant content loss detected. Skipping validation report update.")
        return True

    try:
        # Load existing workbook or create new one
        if validation_report_path.exists():
            wb = load_workbook(validation_report_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Validation Report"

            # Add headers
            headers = ['File Name', 'Issue Type', 'Severity', 'Description', 'Details']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

        ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        # Add content loss entry
        ws.cell(row=next_row, column=1, value=epub_name)
        ws.cell(row=next_row, column=2, value='Content Loss')

        # Determine severity
        if result.content_loss_percentage > 10:
            severity = 'HIGH'
            severity_color = "FF6B6B"
        elif result.content_loss_percentage > 5:
            severity = 'MEDIUM'
            severity_color = "FFE66D"
        else:
            severity = 'LOW'
            severity_color = "FFD93D"

        severity_cell = ws.cell(row=next_row, column=3, value=severity)
        severity_cell.fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type="solid")

        description = f"Content loss detected: {result.content_loss_percentage:.2f}% ({result.total_epub_chars - result.total_xml_chars} characters)"
        ws.cell(row=next_row, column=4, value=description)

        details = f"EPUB: {result.total_epub_chars} chars, XML: {result.total_xml_chars} chars, Differences: {len(result.differences)}"
        if result.missing_chapters:
            details += f", Missing chapters: {', '.join(result.missing_chapters)}"
        ws.cell(row=next_row, column=5, value=details)

        # Adjust column widths if this is a new file
        if next_row == 2:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 80

        # Save
        wb.save(validation_report_path)
        logger.info(f"Added content loss entry to validation report: {validation_report_path}")
        return True

    except Exception as e:
        logger.error(f"Error updating validation report: {e}")
        return False


def run_content_diff(epub_path: Path, rittdoc_path: Path, output_dir: Path,
                     validation_report_path: Optional[Path] = None,
                     skip_detailed_diff: bool = False) -> ContentDiffResult:
    """
    Run content diff comparison and generate reports.

    Args:
        epub_path: Path to original EPUB file
        rittdoc_path: Path to converted rittdoc zip file
        output_dir: Directory for output files
        validation_report_path: Optional path to validation report Excel
        skip_detailed_diff: If True, skip detailed diff for faster processing

    Returns:
        ContentDiffResult with comparison results
    """
    # Ensure output directory exists
    output_dir.mkdir(parents=True, exist_ok=True)

    # Run comparison
    result = compare_content(epub_path, rittdoc_path, skip_detailed_diff=skip_detailed_diff)

    # Write content diff Excel
    contentdiff_path = output_dir / "contentdiff.xlsx"
    write_contentdiff_excel(result, contentdiff_path)

    # Add to validation report if content loss detected
    if validation_report_path and result.content_loss_detected:
        add_to_validation_report(validation_report_path, result, epub_path.name)
    elif result.content_loss_detected and not validation_report_path:
        # Default validation report location
        default_report = output_dir / "validation_report.xlsx"
        add_to_validation_report(default_report, result, epub_path.name)

    return result


def main():
    """Command-line interface for content diff tool."""
    import argparse

    parser = argparse.ArgumentParser(description='Compare EPUB and rittdoc content')
    parser.add_argument('epub', type=Path, help='Path to original EPUB file')
    parser.add_argument('rittdoc', type=Path, help='Path to rittdoc zip file')
    parser.add_argument('-o', '--output', type=Path, default=Path('.'),
                       help='Output directory for reports')
    parser.add_argument('-v', '--validation-report', type=Path,
                       help='Path to validation report Excel file')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    parser.add_argument('--fast', action='store_true',
                       help='Fast mode: skip detailed diff comparison (much faster for large files)')

    args = parser.parse_args()

    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    # Run comparison
    result = run_content_diff(args.epub, args.rittdoc, args.output, args.validation_report,
                              skip_detailed_diff=args.fast)

    # Print summary
    print("\n" + "=" * 60)
    print("CONTENT DIFF SUMMARY")
    print("=" * 60)
    print(f"EPUB Characters: {result.total_epub_chars:,}")
    print(f"XML Characters:  {result.total_xml_chars:,}")
    print(f"Difference:      {result.total_epub_chars - result.total_xml_chars:,}")
    print(f"Content Loss:    {result.content_loss_percentage:.2f}%")
    print(f"Loss Detected:   {'YES' if result.content_loss_detected else 'NO'}")
    print(f"Differences:     {len(result.differences)}")
    if result.missing_chapters:
        print(f"Missing Chapters: {', '.join(result.missing_chapters)}")
    print("=" * 60)

    return 0 if not result.content_loss_detected else 1


if __name__ == '__main__':
    exit(main())
